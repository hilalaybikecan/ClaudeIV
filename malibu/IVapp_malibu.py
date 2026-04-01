import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
import os
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import seaborn as sns
import re
import json
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class IVDataAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("IV Data Analyzer")
        self.root.geometry("1200x800")
        
        # Main data storage
        self.measurements_data = pd.DataFrame(columns=[
            'Filename', 'Substrate ID', 'Pixel', 'Scan Direction',
            'Voc [V]', 'Jsc [mA/cm2]', 'FF [.]', 'Efficiency [.]',
            'Pmpp [W/m2]', 'Vmpp [V]', 'Jmpp [mA/cm2]', 'Roc [Ohm.m2]', 'Rsc [Ohm.m2]', 'Scan Speed [V/s]', 'Cell Area [cm2]', 'Filepath'
        ])
        
        # Persistent copy for display sorting (never modified directly)
        self.measurements_data_original = pd.DataFrame(columns=[
            'Filename', 'Substrate ID', 'Pixel', 'Scan Direction', 
            'Voc [V]', 'Jsc [mA/cm2]', 'FF [.]', 'Efficiency [.]', 
            'Pmpp [W/m2]', 'Vmpp [V]', 'Jmpp [mA/cm2]', 'Roc [Ohm.m2]', 'Rsc [Ohm.m2]'
        ])
        
        # Conditions data storage
        self.conditions_data = pd.DataFrame(columns=['Substrate ID', 'Condition', 'Display Order'])
        
        # Store the current figure for saving
        self.current_figure = None
        self.current_plot_data = None
        
        # Track sort state for measurements tree
        self.last_sorted_column = None
        self.last_sort_descending = False
        
        # Available color palettes
        self.color_palettes = {
            "muted": "muted",
            "deep": "deep",
            "colorblind": "colorblind",
            "pastel": "pastel",
            "bright": "bright",
            "dark": "dark",
            "tab10": "tab10",
            "tab20": "tab20",
            "Set1": "Set1",
            "Set2": "Set2",
            "Set3": "Set3",
            "Paired": "Paired",
            "viridis": "viridis",
            "magma": "magma",
            "plasma": "plasma",
            "inferno": "inferno",
            "rocket": "rocket",
            "mako": "mako",
            "flare": "flare",
            "crest": "crest",
            "Spectral": "Spectral",
            "RdYlGn": "RdYlGn",
            "coolwarm": "coolwarm",
            "husl": "husl",
            "Default (blue)": None,
        }
        
        # Auto-save custom order to temp file
        self.auto_save_file = "temp_condition_order.json"
        
        # Create UI elements
        self.create_ui()
    
    def create_ui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
        # Existing tabs
        data_frame = ttk.Frame(notebook)
        notebook.add(data_frame, text="Data Management")
    
        plot_frame = ttk.Frame(notebook)
        notebook.add(plot_frame, text="Condition Plots")
    
        # New IV Plot tab
        iv_plot_frame = ttk.Frame(notebook)
        notebook.add(iv_plot_frame, text="IV Plots")

        # Best PCE tab
        best_pce_frame = ttk.Frame(notebook)
        notebook.add(best_pce_frame, text="Best PCE")

        # Setup UI components
        self.setup_data_management(data_frame)
        self.setup_plotting(plot_frame)
        self.setup_iv_plot(iv_plot_frame)
        self.setup_best_pce_tab(best_pce_frame)

    
    def setup_data_management(self, parent):
        # File management frame
        file_frame = ttk.LabelFrame(parent, text="File Management")
        file_frame.pack(fill="x", expand=False, padx=10, pady=5)
        
        # Browse button
        browse_button = ttk.Button(file_frame, text="Browse for Files", command=self.browse_files)
        browse_button.pack(side="left", padx=5, pady=5)
        
        # Remove button
        remove_button = ttk.Button(file_frame, text="Remove Selected", command=self.remove_selected)
        remove_button.pack(side="left", padx=5, pady=5)
        
        # Measurements table frame
        measurements_frame = ttk.LabelFrame(parent, text="Measurements")
        measurements_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Create treeview for measurements data
        columns = ('Filename', 'Substrate ID', 'Pixel', 'Scan Direction',
                  'Voc [V]', 'Jsc [mA/cm2]', 'FF [.]', 'Efficiency [.]',
                  'Pmpp [W/m2]', 'Vmpp [V]', 'Jmpp [mA/cm2]', 'Roc [Ohm.m2]', 'Rsc [Ohm.m2]', 'Scan Speed [V/s]', 'Cell Area [cm2]')
        self.display_columns = columns
        
        self.measurements_tree = ttk.Treeview(measurements_frame, columns=self.display_columns, show='headings')
        
        # Set column headings and widths
        for col in self.display_columns:
            self.measurements_tree.heading(col, text=col, command=lambda c=col: self.sort_measurements_by_column(c))
            width = 80 if col != 'Filename' else 200  # Set wider for filename
            self.measurements_tree.column(col, width=width, anchor='center')
        
        # Add scrollbars
        vsb = ttk.Scrollbar(measurements_frame, orient="vertical", command=self.measurements_tree.yview)
        hsb = ttk.Scrollbar(measurements_frame, orient="horizontal", command=self.measurements_tree.xview)
        self.measurements_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Pack scrollbars and treeview
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.measurements_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Conditions frame
        conditions_frame = ttk.LabelFrame(parent, text="Conditions")
        conditions_frame.pack(fill="both", expand=False, padx=10, pady=5)
        
        # Conditions management
        conditions_input_frame = ttk.Frame(conditions_frame)
        conditions_input_frame.pack(fill="x", expand=False, padx=5, pady=5)
        
        ttk.Label(conditions_input_frame, text="New Condition:").pack(side="left", padx=5)
        self.condition_entry = ttk.Entry(conditions_input_frame, width=30)
        self.condition_entry.pack(side="left", padx=5)
        
        add_condition_button = ttk.Button(conditions_input_frame, text="Add Condition", command=self.add_condition)
        add_condition_button.pack(side="left", padx=5)
        
        # Conditions assignment frame
        assignment_frame = ttk.Frame(conditions_frame)
        assignment_frame.pack(fill="x", expand=False, padx=5, pady=5)
        
        ttk.Label(assignment_frame, text="Assign selected measurements to:").pack(side="left", padx=5)
        self.condition_combobox = ttk.Combobox(assignment_frame, width=30, state="readonly")
        self.condition_combobox.pack(side="left", padx=5)
        
        assign_button = ttk.Button(assignment_frame, text="Assign", command=self.assign_condition)
        assign_button.pack(side="left", padx=5)
        
        # Save/Load conditions frame
        save_load_frame = ttk.Frame(conditions_frame)
        save_load_frame.pack(fill="x", expand=False, padx=5, pady=5)
        
        save_conditions_button = ttk.Button(save_load_frame, text="Save Conditions", command=self.save_conditions)
        save_conditions_button.pack(side="left", padx=5)
        
        load_conditions_button = ttk.Button(save_load_frame, text="Load Conditions", command=self.load_conditions)
        load_conditions_button.pack(side="left", padx=5)
        
        # Conditions table frame
        conditions_table_frame = ttk.Frame(conditions_frame)
        conditions_table_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create treeview for conditions with order column
        self.conditions_tree = ttk.Treeview(
            conditions_table_frame, 
            columns=('Substrate ID', 'Condition', 'Display Order'), 
            show='headings'
        )
        self.conditions_tree.heading('Substrate ID', text='Substrate ID')
        self.conditions_tree.heading('Condition', text='Condition (double-click to rename)')
        self.conditions_tree.heading('Display Order', text='Display Order')
        self.conditions_tree.column('Substrate ID', width=100, anchor='center')
        self.conditions_tree.column('Condition', width=200, anchor='center')
        self.conditions_tree.bind('<Double-1>', self.on_condition_tree_double_click)
        self.conditions_tree.column('Display Order', width=80, anchor='center')
        
        # Add summary frame for unique conditions
        summary_frame = ttk.LabelFrame(conditions_frame, text="Condition Summary")
        summary_frame.pack(fill="x", expand=False, padx=10, pady=(0, 5))
        
        self.summary_label = ttk.Label(summary_frame, text="No conditions assigned yet")
        self.summary_label.pack(padx=5, pady=5)
        
        # Add scrollbar
        conditions_vsb = ttk.Scrollbar(conditions_table_frame, orient="vertical", command=self.conditions_tree.yview)
        self.conditions_tree.configure(yscrollcommand=conditions_vsb.set)
        
        # Pack scrollbar and treeview
        conditions_vsb.pack(side="right", fill="y")
        self.conditions_tree.pack(fill="both", expand=True)
        
        # Buttons for reordering conditions
        conditions_order_frame = ttk.Frame(conditions_frame)
        conditions_order_frame.pack(fill="x", expand=False, padx=5, pady=5)
        
        move_up_button = ttk.Button(conditions_order_frame, text="Move Up", command=lambda: self.reorder_condition(-1))
        move_up_button.pack(side="left", padx=5)
        
        move_down_button = ttk.Button(conditions_order_frame, text="Move Down", command=lambda: self.reorder_condition(1))
        move_down_button.pack(side="left", padx=5)
    
    def setup_plotting(self, parent):
        # Parameters for plotting
        self.plot_parameters = ['Voc [V]', 'Jsc [mA/cm2]', 'FF [.]', 'Efficiency [.]', 'Pmpp [W/m2]', 'Vmpp [V]', 'Jmpp [mA/cm2]', 'Roc [Ohm.m2]', 'Rsc [Ohm.m2]', 'Scan Speed [V/s]']

        # Horizontal container: left = plot controls, right = condition panels
        controls_container = ttk.Frame(parent)
        controls_container.pack(fill="x", expand=False, padx=10, pady=5)

        # ── LEFT: Plot controls ──────────────────────────────────────────────
        plot_control_frame = ttk.LabelFrame(controls_container, text="Plot Controls")
        plot_control_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Multi-plot management
        multiplot_frame = ttk.Frame(plot_control_frame)
        multiplot_frame.pack(fill="x", expand=False, padx=5, pady=5)

        ttk.Label(multiplot_frame, text="Plots:").pack(side="left", padx=5)

        add_plot_btn = ttk.Button(multiplot_frame, text="Add Plot", command=self.add_plot_slot)
        add_plot_btn.pack(side="left", padx=5)

        remove_plot_btn = ttk.Button(multiplot_frame, text="Remove Last", command=self.remove_plot_slot)
        remove_plot_btn.pack(side="left", padx=5)

        # Container for plot parameter selections
        self.plot_slots_frame = ttk.Frame(plot_control_frame)
        self.plot_slots_frame.pack(fill="x", expand=False, padx=5, pady=5)

        # Store plot slot widgets
        self.plot_slots = []

        # Add first plot by default
        self.add_plot_slot()

        # Color palette selection
        color_frame = ttk.Frame(plot_control_frame)
        color_frame.pack(fill="x", expand=False, padx=5, pady=5)

        ttk.Label(color_frame, text="Color Palette:").pack(side="left", padx=5)

        self.color_palette_combobox = ttk.Combobox(
            color_frame,
            values=list(self.color_palettes.keys()),
            width=20,
            state="readonly"
        )
        self.color_palette_combobox.pack(side="left", padx=5)
        self.color_palette_combobox.current(0)  # Select muted palette as default
        self.color_palette_combobox.bind("<<ComboboxSelected>>", self.on_color_mode_change)

        # Manual color selection button
        self.manual_colors_btn = ttk.Button(color_frame, text="Manual Colors", command=self.open_manual_color_picker)
        self.manual_colors_btn.pack(side="left", padx=5)

        # Store manual color selections
        self.manual_colors = {}  # {condition_name: color_hex}

        # Tick label size controls
        tick_frame = ttk.Frame(plot_control_frame)
        tick_frame.pack(fill="x", expand=False, padx=5, pady=5)

        ttk.Label(tick_frame, text="X-tick size:").pack(side="left", padx=5)
        self.xtick_size_var = tk.StringVar(value="16")
        ttk.Spinbox(tick_frame, from_=6, to=40, increment=1, textvariable=self.xtick_size_var, width=5).pack(side="left", padx=(0, 15))

        ttk.Label(tick_frame, text="Y-tick size:").pack(side="left", padx=5)
        self.ytick_size_var = tk.StringVar(value="10")
        ttk.Spinbox(tick_frame, from_=6, to=40, increment=1, textvariable=self.ytick_size_var, width=5).pack(side="left", padx=(0, 15))

        ttk.Label(tick_frame, text="Y-label size:").pack(side="left", padx=5)
        self.ylabel_size_var = tk.StringVar(value="20")
        ttk.Spinbox(tick_frame, from_=6, to=40, increment=1, textvariable=self.ylabel_size_var, width=5).pack(side="left", padx=(0, 15))

        ttk.Label(tick_frame, text="X-tick angle:").pack(side="left", padx=5)
        self.xtick_angle_var = tk.StringVar(value="45")
        ttk.Spinbox(tick_frame, from_=0, to=90, increment=15, textvariable=self.xtick_angle_var, width=5).pack(side="left", padx=0)

        # X-axis ordering selection
        xorder_frame = ttk.Frame(plot_control_frame)
        xorder_frame.pack(fill="x", expand=False, padx=5, pady=5)

        ttk.Label(xorder_frame, text="X-axis Order:").pack(side="left", padx=5)

        self.xorder_combobox = ttk.Combobox(
            xorder_frame,
            values=["Display Order (Default)", "Alphabetical", "Efficiency (high→low)", "Custom"],
            width=25,
            state="readonly"
        )
        self.xorder_combobox.pack(side="left", padx=5)
        self.xorder_combobox.current(0)
        self.xorder_combobox.bind("<<ComboboxSelected>>", self.on_xorder_change)

        # Plot by Substrate ID mode (skip condition assignment)
        substrate_mode_frame = ttk.Frame(plot_control_frame)
        substrate_mode_frame.pack(fill="x", expand=False, padx=5, pady=5)

        self.plot_by_substrate_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(substrate_mode_frame, text="Plot by Substrate ID (each sample is its own condition)",
                        variable=self.plot_by_substrate_var,
                        command=self.on_substrate_mode_change).pack(side="left", padx=5)

        # Best measurement filter
        best_meas_frame = ttk.Frame(plot_control_frame)
        best_meas_frame.pack(fill="x", expand=False, padx=5, pady=5)

        self.best_measurement_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(best_meas_frame, text="Best measurement only (highest avg. efficiency fwd/rev pair per cell)",
                        variable=self.best_measurement_var).pack(side="left", padx=5)

        # Plot buttons
        button_frame = ttk.Frame(plot_control_frame)
        button_frame.pack(fill="x", expand=False, padx=5, pady=5)

        ttk.Button(button_frame, text="Generate Plot", command=self.generate_plot).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Save Plot", command=self.save_plot).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Export Data to Excel", command=self.export_condition_data).pack(side="left", padx=5)

        # ── RIGHT: Condition panels ──────────────────────────────────────────
        right_panel = ttk.Frame(controls_container)
        right_panel.pack(side="left", fill="both", padx=(5, 0))

        # Conditions to Plot filter
        cond_filter_frame = ttk.LabelFrame(right_panel, text="Conditions to Plot")
        cond_filter_frame.pack(fill="both", expand=True, pady=(0, 5))

        cond_filter_inner = ttk.Frame(cond_filter_frame)
        cond_filter_inner.pack(fill="both", expand=True, padx=5, pady=5)

        cond_list_scroll = ttk.Frame(cond_filter_inner)
        cond_list_scroll.pack(side="left", fill="both", expand=True)

        self.plot_cond_listbox = tk.Listbox(cond_list_scroll, selectmode=tk.MULTIPLE, height=6, exportselection=False, width=22)
        cond_vsb = ttk.Scrollbar(cond_list_scroll, orient="vertical", command=self.plot_cond_listbox.yview)
        self.plot_cond_listbox.configure(yscrollcommand=cond_vsb.set)
        self.plot_cond_listbox.pack(side="left", fill="both", expand=True)
        cond_vsb.pack(side="right", fill="y")

        cond_btn_frame = ttk.Frame(cond_filter_inner)
        cond_btn_frame.pack(side="left", padx=(5, 0), fill="y")
        ttk.Button(cond_btn_frame, text="All", command=self._select_all_plot_conds).pack(fill="x", pady=2)
        ttk.Button(cond_btn_frame, text="None", command=self._deselect_all_plot_conds).pack(fill="x", pady=2)

        # Custom order interface (initially hidden, shown in right_panel when "Custom" selected)
        self.custom_order_frame = ttk.Frame(right_panel)

        custom_order_labelframe = ttk.LabelFrame(self.custom_order_frame, text="Custom Condition Order")
        custom_order_labelframe.pack(fill="both", expand=True, padx=0, pady=0)

        listbox_frame = ttk.Frame(custom_order_labelframe)
        listbox_frame.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Label(listbox_frame, text="Ctrl+Up/Down or double-click to move to top:").pack(anchor="w", pady=(0, 3))

        list_scroll_frame = ttk.Frame(listbox_frame)
        list_scroll_frame.pack(fill="both", expand=True)

        self.conditions_listbox = tk.Listbox(list_scroll_frame, height=5, selectmode=tk.SINGLE, width=22)
        self.conditions_listbox.pack(side="left", fill="both", expand=True)

        self.conditions_listbox.bind('<Control-Up>', lambda e: self.move_condition_up())
        self.conditions_listbox.bind('<Control-Down>', lambda e: self.move_condition_down())
        self.conditions_listbox.bind('<Double-Button-1>', self.on_condition_double_click)

        listbox_scrollbar = ttk.Scrollbar(list_scroll_frame, orient="vertical", command=self.conditions_listbox.yview)
        listbox_scrollbar.pack(side="right", fill="y")
        self.conditions_listbox.configure(yscrollcommand=listbox_scrollbar.set)

        buttons_frame = ttk.Frame(listbox_frame)
        buttons_frame.pack(fill="x", pady=(3, 0))

        self.move_up_btn = ttk.Button(buttons_frame, text="↑", command=self.move_condition_up, width=3)
        self.move_up_btn.pack(side="left", padx=(0, 2))
        self.move_down_btn = ttk.Button(buttons_frame, text="↓", command=self.move_condition_down, width=3)
        self.move_down_btn.pack(side="left", padx=2)
        self.reset_order_btn = ttk.Button(buttons_frame, text="Reset", command=self.reset_condition_order)
        self.reset_order_btn.pack(side="left", padx=2)

        # Frame for the condition plot
        self.condition_plot_frame = ttk.Frame(parent)
        self.condition_plot_frame.pack(fill="both", expand=True, padx=10, pady=5)

    def add_plot_slot(self):
        """Add a new plot slot (max 4)"""
        if len(self.plot_slots) >= 4:
            messagebox.showwarning("Maximum Reached", "You can add a maximum of 4 plots.")
            return

        slot_frame = ttk.Frame(self.plot_slots_frame)
        slot_frame.pack(fill="x", pady=2)

        ttk.Label(slot_frame, text=f"Plot {len(self.plot_slots) + 1}:").pack(side="left", padx=5)

        param_combo = ttk.Combobox(slot_frame, values=self.plot_parameters, width=20, state="readonly")
        param_combo.pack(side="left", padx=5)
        param_combo.current(len(self.plot_slots) % len(self.plot_parameters))  # Different default for each

        # Y-axis limits for this plot
        ttk.Label(slot_frame, text="Y-min:").pack(side="left", padx=(10, 2))
        ymin_entry = ttk.Entry(slot_frame, width=8)
        ymin_entry.pack(side="left", padx=2)

        ttk.Label(slot_frame, text="Y-max:").pack(side="left", padx=(5, 2))
        ymax_entry = ttk.Entry(slot_frame, width=8)
        ymax_entry.pack(side="left", padx=2)

        self.plot_slots.append({
            'frame': slot_frame,
            'combobox': param_combo,
            'ymin': ymin_entry,
            'ymax': ymax_entry
        })

    def remove_plot_slot(self):
        """Remove the last plot slot"""
        if len(self.plot_slots) <= 1:
            messagebox.showwarning("Minimum Required", "At least one plot is required.")
            return

        last_slot = self.plot_slots.pop()
        last_slot['frame'].destroy()
    
    def save_conditions(self):
        """Save conditions data to a JSON file"""
        if self.conditions_data.empty:
            messagebox.showwarning("Warning", "No conditions to save.")
            return
        
        # Get conditions to save
        custom_order = [self.conditions_listbox.get(i) for i in range(self.conditions_listbox.size())] if hasattr(self, 'conditions_listbox') else []
        conditions_to_save = {
            "conditions_list": list(self.condition_combobox['values']),
            "conditions_data": self.conditions_data.to_dict(orient='records'),
            "custom_order": custom_order
        }
        
        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save Conditions As"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            with open(file_path, 'w') as f:
                json.dump(conditions_to_save, f, indent=4)
            messagebox.showinfo("Success", f"Conditions saved successfully to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save conditions: {str(e)}")
    
    def load_conditions(self):
        """Load conditions data from a JSON file"""
        # Ask for file to load
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Load Conditions"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            with open(file_path, 'r') as f:
                saved_data = json.load(f)
            
            # Load conditions list for dropdown
            self.condition_combobox['values'] = saved_data.get("conditions_list", [])
            
            # Load conditions data
            conditions_data = saved_data.get("conditions_data", [])
            self.conditions_data = pd.DataFrame(conditions_data)
            
            # Update conditions tree
            self.update_conditions_tree()
            
            # Load custom order if available
            custom_order = saved_data.get("custom_order", [])
            if custom_order and hasattr(self, 'conditions_listbox'):
                self.populate_conditions_listbox(custom_order)
                
            messagebox.showinfo("Success", f"Conditions loaded successfully from:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load conditions: {str(e)}")
    
    def browse_files(self):
        filepaths = filedialog.askopenfilenames(
            title="Select IV Measurement Files",
            filetypes=(("IV Files", "*.iv"), ("All Files", "*.*"))
        )
        
        if not filepaths:
            return
        
        # Collect already-loaded filenames to skip duplicates
        existing_filenames = set(self.measurements_data['Filename'].tolist()) if not self.measurements_data.empty else set()

        for filepath in filepaths:
            filename = os.path.basename(filepath)
            if filename in existing_filenames:
                continue  # Skip already-loaded file

            # Extract data from file
            try:
                data = self.parse_iv_file(filepath)

                # Add to both dataframes
                # Align columns in case of new fields
                for col in data.keys():
                    if col not in self.measurements_data.columns:
                        self.measurements_data[col] = None
                self.measurements_data = pd.concat([self.measurements_data, pd.DataFrame([data])], ignore_index=True)
                self.measurements_data_original = self.measurements_data.copy()
                existing_filenames.add(filename)

            except Exception as e:
                messagebox.showerror("Error", f"Failed to parse file {os.path.basename(filepath)}: {str(e)}")
        
        # Deduplicate measurements by Filename
        self.measurements_data = self.measurements_data.drop_duplicates(subset='Filename', keep='first').reset_index(drop=True)
        self.measurements_data_original = self.measurements_data.copy()

        self.update_measurements_display()
        self.refresh_iv_selection()

    def parse_iv_file(self, filepath):
        filename = os.path.basename(filepath)
        
        # Extract pixel information from filename

        pixel_match = re.search(r'c(\d)', filename)
        pixel = {'1': 'A', '2': 'B', '3': 'C'}.get(pixel_match.group(1), 'Unknown') if pixel_match else 'Unknown'
        
        with open(filepath, 'r') as f:
            content = f.read()

        # Extract sweep metadata to compute scan speed (V/s)
        vstart = self.extract_value(content, r'Vstart:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        vend = self.extract_value(content, r'Vend:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        npoints = None
        m_np = re.search(r'Number of points:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)', content)
        if m_np:
            try:
                npoints = int(float(m_np.group(1)))
            except:
                npoints = None
        delay_s = self.extract_value(content, r'Delay \[s\]:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        it_s = self.extract_value(content, r'Integration time \[s\]:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        scan_speed = None
        try:
            if vstart is not None and vend is not None and npoints and npoints > 1 and (delay_s is not None) and (it_s is not None):
                step_time = (delay_s + it_s)
                total_time = (npoints - 1) * step_time
                if total_time > 0:
                    scan_speed = abs(vend - vstart) / total_time
        except Exception:
            scan_speed = None

        
        # Extract substrate ID
        substrate_id_match = re.search(r'Deposition ID:\s*([A-Za-z0-9_]+)', content)
        substrate_id = substrate_id_match.group(1) if substrate_id_match else 'Unknown'

        # Extract cell area (in m2, convert to cm2)
        cell_area_cm2 = None
        area_match = re.search(r'Cell size \[m2\]:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)', content)
        if area_match:
            try:
                area_m2 = float(area_match.group(1))
                cell_area_cm2 = area_m2 * 10000  # Convert m² to cm²
            except:
                cell_area_cm2 = None
        
        # Extract measurement data
        data_section_match = re.search(r'% MEASURED IV FRLOOP DATA.*?\nV \(measured\) \[V\].*?\n(.*?)$', content, re.DOTALL)
        if not data_section_match:
            raise ValueError("Could not find measurement data section in the file.")
        
        data_lines = data_section_match.group(1).strip().split('\n')
        voltages = []
        
        for line in data_lines:
            if line:
                values = line.split('\t')
                if len(values) >= 2:
                    try:
                        voltages.append(float(values[0]))
                    except:
                        pass
        
        # Determine scan direction
        if len(voltages) > 1:
            scan_direction = "fwd" if voltages[1] > voltages[0] else "rev"
        else:
            scan_direction = "unknown"
        
        # Extract analysis outputs
        analysis_outputs = {
            'Voc [V]': self.extract_value(content, r'Voc \[V\]:\s*(\d+\.\d+E[+-]\d+)'),
            'Jsc [mA/cm2]': self.extract_value(content, r'Jsc \[A/m2\]:\s*(\d+\.\d+E[+-]\d+)', convert_to_ma_cm2=True),
            'FF [.]': self.extract_value(content, r'FF \[.\]:\s*(\d+\.\d+E[+-]\d+)'),
            'Efficiency [.]': self.extract_value(content, r'Efficiency \[.\]:\s*(\d+\.\d+E[+-]\d+)'),
            'Pmpp [W/m2]': self.extract_value(content, r'Pmpp \[W/m2\]:\s*(\d+\.\d+E[+-]\d+)'),
            'Vmpp [V]': self.extract_value(content, r'Vmpp \[V\]:\s*(\d+\.\d+E[+-]\d+)'),
            'Jmpp [mA/cm2]': self.extract_value(content, r'Jmpp \[A\]:\s*(\d+\.\d+E[+-]\d+)', convert_to_ma_cm2=True),
            'Roc [Ohm.m2]': self.extract_value(content, r'Roc \[Ohm\.m2\]:\s*(\d+\.\d+E[+-]\d+)'),
            'Rsc [Ohm.m2]': self.extract_value(content, r'Rsc \[Ohm\.m2\]:\s*(\d+\.\d+E[+-]\d+)')
        }
        
        # Create data dictionary
        data = {
            'Filename': filename,
            'Substrate ID': substrate_id,
            'Pixel': pixel,
            'Scan Direction': scan_direction,
            'Scan Speed [V/s]': scan_speed,
            'Cell Area [cm2]': cell_area_cm2,
            'Filepath': filepath,
            **analysis_outputs
        }
        
        return data
    
    def extract_value(self, content, pattern, convert_to_ma_cm2=False):
        match = re.search(pattern, content)
        if match:
            try:
                value = float(match.group(1))
                if "Efficiency" in pattern or "FF" in pattern:
                    return 100 * value
                elif convert_to_ma_cm2:
                    return value * 0.1  # Convert A/m2 to mA/cm2
                else:
                    return value
            except:
                return None
        return None
    
    def update_measurements_display(self):
        # Clear treeview
        for item in self.measurements_tree.get_children():
            self.measurements_tree.delete(item)
        
        # Repopulate with sorted data
        for _, row in self.measurements_data.iterrows():
            values = [row.get(col, '') for col in self.display_columns]
            self.measurements_tree.insert('', 'end', values=values)
    
    def remove_selected(self):
        selected_items = self.measurements_tree.selection()
        if not selected_items:
            return
        
        for item in selected_items:
            item_values = self.measurements_tree.item(item, 'values')
            filename = item_values[0]
            
            # Remove from both dataframes
            self.measurements_data = self.measurements_data[self.measurements_data['Filename'] != filename]
            self.measurements_data_original = self.measurements_data.copy()
            
            # Remove from treeview
            self.measurements_tree.delete(item)
    
    def add_condition(self):
        condition = self.condition_entry.get().strip()
        if not condition:
            messagebox.showwarning("Warning", "Please enter a condition name.")
            return
        
        # Check if condition already exists
        existing_conditions = self.get_conditions()
        if condition in existing_conditions:
            messagebox.showwarning("Warning", f"Condition '{condition}' already exists.")
            return
        
        # Add condition to combobox
        existing_conditions.append(condition)
        self.condition_combobox['values'] = existing_conditions
        self.condition_entry.delete(0, tk.END)
    
    def get_conditions(self):
        return list(self.condition_combobox['values']) if self.condition_combobox['values'] else []
    
    def assign_condition(self):
        selected_items = self.measurements_tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select one or more measurements.")
            return
        
        condition = self.condition_combobox.get()
        if not condition:
            messagebox.showwarning("Warning", "Please select a condition.")
            return
        
        # Get unique substrate IDs from selected items
        substrate_ids = set()
        for item in selected_items:
            item_values = self.measurements_tree.item(item, 'values')
            substrate_id = item_values[1]  # Substrate ID is the second column
            substrate_ids.add(substrate_id)
        
        # Determine next display order value
        next_order = 1
        if not self.conditions_data.empty:
            next_order = self.conditions_data['Display Order'].max() + 1 if not self.conditions_data.empty else 1
        
        # Update conditions data
        for substrate_id in substrate_ids:
            # Remove existing condition for this substrate
            self.conditions_data = self.conditions_data[self.conditions_data['Substrate ID'] != substrate_id]
            
            # Add new condition
            new_condition = pd.DataFrame({
                'Substrate ID': [substrate_id],
                'Condition': [condition],
                'Display Order': [next_order]
            })
            self.conditions_data = pd.concat([self.conditions_data, new_condition], ignore_index=True)
        
        # Update conditions tree
        self.update_conditions_tree()
    
    def update_conditions_tree(self):
        # Clear tree
        for item in self.conditions_tree.get_children():
            self.conditions_tree.delete(item)
        
        # Sort by display order
        sorted_conditions = self.conditions_data.sort_values(by='Display Order')
        
        # Add conditions to tree
        for _, row in sorted_conditions.iterrows():
            self.conditions_tree.insert('', 'end', values=(row['Substrate ID'], row['Condition'], row['Display Order']))
        
        # Update summary and condition filter listbox
        self.update_condition_summary()
        self.update_plot_conditions_listbox()
    
    def update_condition_summary(self):
        """Update the condition summary display"""
        if self.conditions_data.empty:
            self.summary_label.config(text="No conditions assigned yet")
        else:
            # Get unique conditions and their counts
            condition_counts = self.conditions_data['Condition'].value_counts().sort_index()
            summary_text = "Unique conditions: " + ", ".join([f"{cond} ({count})" for cond, count in condition_counts.items()])
            self.summary_label.config(text=summary_text)
    
    def reorder_condition(self, direction):
        selected_items = self.conditions_tree.selection()
        if not selected_items or len(selected_items) != 1:
            messagebox.showwarning("Warning", "Please select exactly one condition to reorder.")
            return
        
        item = selected_items[0]
        item_values = self.conditions_tree.item(item, 'values')
        substrate_id = item_values[0]
        condition = item_values[1]
        current_order = int(item_values[2])
        
        # Find the condition to swap with
        target_order = current_order + direction
        
        # Check if target order is valid
        if target_order < 1 or target_order > len(self.conditions_tree.get_children()):
            return  # Nothing to swap with
        
        # Find the item to swap with
        target_item = None
        for tree_item in self.conditions_tree.get_children():
            tree_item_values = self.conditions_tree.item(tree_item, 'values')
            if int(tree_item_values[2]) == target_order:
                target_item = tree_item
                target_substrate_id = tree_item_values[0]
                target_condition = tree_item_values[1]
                break
        
        if target_item:
            # Swap orders in the dataframe
            self.conditions_data.loc[
                self.conditions_data['Substrate ID'] == substrate_id, 'Display Order'
            ] = target_order
            
            self.conditions_data.loc[
                self.conditions_data['Substrate ID'] == target_substrate_id, 'Display Order'
            ] = current_order
            
            # Update the treeview
            self.update_conditions_tree()
    
    def sort_measurements_by_column(self, col):
        # Determine sort direction
        if self.last_sorted_column == col:
            # Same column clicked - toggle direction
            descending = not self.last_sort_descending
        else:
            # New column clicked - start with ascending
            descending = False
        
        # Update tracking variables
        self.last_sorted_column = col
        self.last_sort_descending = descending
        
        # Create a sorted copy of the original data for display
        try:
            # For numerical columns, ensure proper numeric sorting
            if col not in ['Filename', 'Substrate ID', 'Pixel', 'Scan Direction']:
                # Convert column to numeric, handling any non-numeric values
                self.measurements_data[col] = pd.to_numeric(self.measurements_data[col], errors='coerce')
            
            # Sort the data using pandas (creates a copy, doesn't modify original)
            sorted_data = self.measurements_data_original.sort_values(by=col, ascending=not descending, na_position='last')
            
            # Update the display data (this is what gets shown)
            self.measurements_data = sorted_data.copy()
            
            # Update the display
            self.update_measurements_display()

        except Exception as e:
            # Fallback to string sorting if numeric sorting fails
            sorted_data = self.measurements_data_original.sort_values(by=col, ascending=not descending, na_position='last', key=lambda x: x.astype(str))
            self.measurements_data = sorted_data.copy()
            self.update_measurements_display()

        # Update column headings to show sort direction
        for column in self.measurements_tree['columns']:
            current_text = self.measurements_tree.heading(column)['text']
            clean_text = current_text.replace(" ▲", "").replace(" ▼", "")
            if column == col:
                sort_indicator = " ▼" if descending else " ▲"
                self.measurements_tree.heading(column, text=clean_text + sort_indicator)
            else:
                self.measurements_tree.heading(column, text=clean_text)
    
    def on_substrate_mode_change(self):
        """Update the condition filter listbox when substrate mode is toggled."""
        self.update_plot_conditions_listbox()
        if self.xorder_combobox.get() == "Custom":
            self.populate_conditions_listbox()

    def on_xorder_change(self, event=None):
        """Handle x-axis ordering combobox selection change"""
        selected_order = self.xorder_combobox.get()
        if selected_order == "Custom":
            self.populate_conditions_listbox()
            self.custom_order_frame.pack(fill="both", expand=True)
        else:
            self.custom_order_frame.pack_forget()
    
    def populate_conditions_listbox(self, custom_order=None):
        """Populate the conditions listbox with current conditions"""
        self.conditions_listbox.delete(0, tk.END)

        substrate_mode = hasattr(self, 'plot_by_substrate_var') and self.plot_by_substrate_var.get()

        # Get unique conditions from the data
        if substrate_mode and not self.measurements_data.empty:
            unique_conditions = sorted(self.measurements_data['Substrate ID'].unique())
        elif not self.measurements_data.empty and not self.conditions_data.empty:
            # Get conditions that have associated measurements
            plot_data = self.measurements_data.merge(
                self.conditions_data,
                on='Substrate ID',
                how='inner'
            )

            if plot_data.empty:
                return
            unique_conditions = plot_data['Condition'].unique()
        else:
            return

        if custom_order:
            # Use provided custom order, then add any missing conditions
            for condition in custom_order:
                if condition in unique_conditions:
                    self.conditions_listbox.insert(tk.END, condition)
            # Add any remaining conditions not in custom order
            for condition in unique_conditions:
                if condition not in custom_order:
                    self.conditions_listbox.insert(tk.END, condition)
        else:
            # Try to auto-load saved custom order
            auto_order = self.auto_load_custom_order()
            if auto_order:
                for condition in auto_order:
                    if condition in unique_conditions:
                        self.conditions_listbox.insert(tk.END, condition)
                for condition in unique_conditions:
                    if condition not in auto_order:
                        self.conditions_listbox.insert(tk.END, condition)
            elif not substrate_mode and not self.conditions_data.empty:
                # Start with default order (Display Order)
                ordered_conditions = self.conditions_data.sort_values(by='Display Order')
                for condition in ordered_conditions['Condition']:
                    if condition in unique_conditions:
                        self.conditions_listbox.insert(tk.END, condition)
            else:
                # Substrate mode or no conditions: use sorted unique_conditions
                for condition in unique_conditions:
                    self.conditions_listbox.insert(tk.END, condition)
    
    def move_condition_up(self):
        """Move selected condition up in the list"""
        selection = self.conditions_listbox.curselection()
        if not selection or selection[0] == 0:
            return  # Nothing selected or already at top
        
        index = selection[0]
        condition = self.conditions_listbox.get(index)
        
        # Remove and reinsert at previous position
        self.conditions_listbox.delete(index)
        self.conditions_listbox.insert(index - 1, condition)
        self.conditions_listbox.selection_set(index - 1)
        self.auto_save_custom_order()
    
    def move_condition_down(self):
        """Move selected condition down in the list"""
        selection = self.conditions_listbox.curselection()
        if not selection or selection[0] == self.conditions_listbox.size() - 1:
            return  # Nothing selected or already at bottom
        
        index = selection[0]
        condition = self.conditions_listbox.get(index)
        
        # Remove and reinsert at next position
        self.conditions_listbox.delete(index)
        self.conditions_listbox.insert(index + 1, condition)
        self.conditions_listbox.selection_set(index + 1)
        self.auto_save_custom_order()
    
    def reset_condition_order(self):
        """Reset the condition order to default (Display Order)"""
        self.populate_conditions_listbox()
    
    def on_condition_double_click(self, event):
        """Move double-clicked condition to the top of the list"""
        selection = self.conditions_listbox.curselection()
        if not selection:
            return
        
        index = selection[0]
        if index == 0:
            return  # Already at top
        
        condition = self.conditions_listbox.get(index)
        self.conditions_listbox.delete(index)
        self.conditions_listbox.insert(0, condition)
        self.conditions_listbox.selection_set(0)
        self.auto_save_custom_order()
    
    def auto_save_custom_order(self):
        """Auto-save current custom order to temporary file"""
        try:
            if hasattr(self, 'conditions_listbox'):
                custom_order = [self.conditions_listbox.get(i) for i in range(self.conditions_listbox.size())]
                with open(self.auto_save_file, 'w') as f:
                    json.dump({"custom_order": custom_order}, f)
        except Exception:
            pass  # Ignore errors in auto-save
    
    def auto_load_custom_order(self):
        """Auto-load custom order from temporary file if it exists"""
        try:
            if os.path.exists(self.auto_save_file):
                with open(self.auto_save_file, 'r') as f:
                    data = json.load(f)
                return data.get("custom_order", [])
        except Exception:
            pass  # Ignore errors in auto-load
        return []
    
    def on_color_mode_change(self, event=None):
        """Handle color palette combobox change"""
        # Reset manual colors if switching to a preset palette
        palette_name = self.color_palette_combobox.get()
        if palette_name != "Manual":
            # Keep manual colors stored but not in use
            pass

    def open_manual_color_picker(self):
        """Open a dialog to manually assign colors to each condition"""
        if self.measurements_data.empty or self.conditions_data.empty:
            messagebox.showwarning("Warning", "Please load data and assign conditions first.")
            return

        # Get unique conditions
        plot_data = self.measurements_data.merge(
            self.conditions_data,
            on='Substrate ID',
            how='inner'
        )

        if plot_data.empty:
            messagebox.showwarning("Warning", "No conditions available for color assignment.")
            return

        unique_conditions = sorted(plot_data['Condition'].unique())

        # Create a new window for color selection
        color_window = tk.Toplevel(self.root)
        color_window.title("Manual Color Selection")
        color_window.geometry("400x500")

        ttk.Label(color_window, text="Assign colors to each condition:", font=('Arial', 12, 'bold')).pack(pady=10)

        # Create a scrollable frame
        canvas = tk.Canvas(color_window)
        scrollbar = ttk.Scrollbar(color_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Store color buttons for each condition
        color_buttons = {}

        def choose_color(condition, btn):
            """Open color chooser and update button"""
            current_color = self.manual_colors.get(condition, "#1f77b4")
            color = colorchooser.askcolor(title=f"Choose color for {condition}", initialcolor=current_color)
            if color[1]:  # color[1] is the hex value
                self.manual_colors[condition] = color[1]
                btn.configure(bg=color[1])

        # Create a row for each condition
        for condition in unique_conditions:
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill="x", padx=10, pady=5)

            # Condition label
            ttk.Label(row_frame, text=condition, width=20).pack(side="left", padx=5)

            # Color preview button
            current_color = self.manual_colors.get(condition, "#1f77b4")
            color_btn = tk.Button(row_frame, text="  ", width=3, bg=current_color,
                                  command=lambda c=condition, b=None: choose_color(c, color_buttons[c]))
            color_btn.pack(side="left", padx=5)
            color_buttons[condition] = color_btn

            # Update the button command with the actual button reference
            color_btn.configure(command=lambda c=condition, b=color_btn: choose_color(c, b))

        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        # Buttons at the bottom
        button_frame = ttk.Frame(color_window)
        button_frame.pack(fill="x", padx=10, pady=10)

        def apply_and_close():
            # Set the palette to use manual colors
            messagebox.showinfo("Success", "Manual colors saved! Generate plot to see the changes.")
            color_window.destroy()

        ttk.Button(button_frame, text="Apply", command=apply_and_close).pack(side="right", padx=5)
        ttk.Button(button_frame, text="Cancel", command=color_window.destroy).pack(side="right", padx=5)

    def on_condition_tree_double_click(self, event):
        """Handle double-click on conditions tree — allow renaming the Condition column."""
        region = self.conditions_tree.identify_region(event.x, event.y)
        if region != 'cell':
            return
        col = self.conditions_tree.identify_column(event.x)
        if col != '#2':  # Only the Condition column
            return
        item = self.conditions_tree.identify_row(event.y)
        if not item:
            return
        old_name = self.conditions_tree.item(item, 'values')[1]
        self._open_rename_dialog(old_name)

    def _open_rename_dialog(self, old_name):
        """Open a small dialog to rename a condition."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Rename Condition")
        dialog.geometry("320x110")
        dialog.resizable(False, False)
        dialog.grab_set()

        ttk.Label(dialog, text=f"Rename  \"{old_name}\"  to:").pack(padx=15, pady=(15, 5))
        name_var = tk.StringVar(value=old_name)
        entry = ttk.Entry(dialog, textvariable=name_var, width=30)
        entry.pack(padx=15)
        entry.select_range(0, tk.END)
        entry.focus_set()

        def apply():
            new_name = name_var.get().strip()
            if not new_name:
                return
            if new_name == old_name:
                dialog.destroy()
                return
            existing = self.get_conditions()
            if new_name in existing:
                messagebox.showwarning("Duplicate", f"Condition '{new_name}' already exists.", parent=dialog)
                return
            self._rename_condition(old_name, new_name)
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=8)
        ttk.Button(btn_frame, text="OK", command=apply, width=8).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=8).pack(side="left", padx=5)
        entry.bind('<Return>', lambda e: apply())
        entry.bind('<Escape>', lambda e: dialog.destroy())

    def _rename_condition(self, old_name, new_name):
        """Apply a condition rename everywhere it appears."""
        # DataFrame
        self.conditions_data['Condition'] = self.conditions_data['Condition'].replace(old_name, new_name)

        # Combobox values
        vals = list(self.condition_combobox['values'])
        vals = [new_name if v == old_name else v for v in vals]
        self.condition_combobox['values'] = vals
        if self.condition_combobox.get() == old_name:
            self.condition_combobox.set(new_name)

        # Custom order listbox
        if hasattr(self, 'conditions_listbox'):
            items = list(self.conditions_listbox.get(0, tk.END))
            self.conditions_listbox.delete(0, tk.END)
            for it in items:
                self.conditions_listbox.insert(tk.END, new_name if it == old_name else it)

        # Conditions to plot listbox
        if hasattr(self, 'plot_cond_listbox'):
            items = list(self.plot_cond_listbox.get(0, tk.END))
            sel = set(self.plot_cond_listbox.curselection())
            self.plot_cond_listbox.delete(0, tk.END)
            for i, it in enumerate(items):
                self.plot_cond_listbox.insert(tk.END, new_name if it == old_name else it)
                if i in sel:
                    self.plot_cond_listbox.selection_set(i)

        # Refresh the tree
        self.update_conditions_tree()

    def update_plot_conditions_listbox(self):
        """Refresh the 'Conditions to Plot' listbox; preserve existing selection if possible."""
        if not hasattr(self, 'plot_cond_listbox'):
            return
        # Remember currently selected conditions
        selected_before = {self.plot_cond_listbox.get(i) for i in self.plot_cond_listbox.curselection()}
        self.plot_cond_listbox.delete(0, tk.END)

        if hasattr(self, 'plot_by_substrate_var') and self.plot_by_substrate_var.get():
            # Substrate ID mode: list all unique substrate IDs
            if self.measurements_data.empty:
                return
            unique_items = self.measurements_data['Substrate ID'].unique()
            for item in sorted(unique_items):
                self.plot_cond_listbox.insert(tk.END, item)
        else:
            # Normal condition mode
            if self.conditions_data.empty:
                return
            unique_items = self.conditions_data.sort_values('Display Order')['Condition'].unique()
            for item in unique_items:
                self.plot_cond_listbox.insert(tk.END, item)

        # Re-select all by default (or restore previous selection)
        for i in range(self.plot_cond_listbox.size()):
            cond = self.plot_cond_listbox.get(i)
            if not selected_before or cond in selected_before:
                self.plot_cond_listbox.selection_set(i)

    def _select_all_plot_conds(self):
        self.plot_cond_listbox.selection_set(0, tk.END)

    def _deselect_all_plot_conds(self):
        self.plot_cond_listbox.selection_clear(0, tk.END)

    def get_condition_order(self, plot_data):
        """Determine the order of conditions for plotting based on user selection"""
        selected_order = self.xorder_combobox.get()
        unique_conditions = plot_data['Condition'].unique()

        if selected_order == "Alphabetical":
            return sorted(unique_conditions)
        elif selected_order == "Efficiency (high→low)":
            # Compute median efficiency per condition, sort descending
            eff_col = 'Efficiency [.]'
            if eff_col in plot_data.columns:
                plot_data[eff_col] = pd.to_numeric(plot_data[eff_col], errors='coerce')
                median_eff = plot_data.groupby('Condition')[eff_col].median().sort_values(ascending=False)
                ordered = [c for c in median_eff.index if c in unique_conditions]
                remaining = [c for c in unique_conditions if c not in ordered]
                return ordered + sorted(remaining)
            return sorted(unique_conditions)
        elif selected_order == "Custom":
            # Get order from listbox
            listbox_order = [self.conditions_listbox.get(i) for i in range(self.conditions_listbox.size())]
            # Filter to only include conditions that exist in the data
            valid_custom_order = [c for c in listbox_order if c in unique_conditions]
            # Add any remaining conditions not specified in custom order
            remaining_conditions = [c for c in unique_conditions if c not in valid_custom_order]
            return valid_custom_order + sorted(remaining_conditions)

        # Default: Display Order (from conditions_data)
        if not self.conditions_data.empty:
            # Sort by display order
            ordered_conditions = self.conditions_data.sort_values(by='Display Order')
            # Filter to only include conditions that exist in the plot data
            return [c for c in ordered_conditions['Condition'].tolist() if c in unique_conditions]
        else:
            # Fallback to alphabetical if no conditions data
            return sorted(unique_conditions)
    
    def _filter_best_measurement(self, df):
        """Keep only the single best fwd/rev pair per Substrate ID.

        For each Substrate ID, all pixels and repeat measurements are
        considered.  Fwd and rev rows within each pixel are sorted by
        filename and paired positionally.  The single pair with the
        highest average Efficiency across all pixels is retained.
        Substrate IDs with only one scan direction are kept as-is.
        """
        kept = []
        for subid, sub_group in df.groupby('Substrate ID'):
            best_avg = -float('inf')
            best_pair = None
            has_any_pair = False

            for pixel, pix_group in sub_group.groupby('Pixel'):
                fwd = pix_group[pix_group['Scan Direction'].str.lower().isin(['fwd', 'fw', 'forward'])].sort_values('Filename')
                rev = pix_group[pix_group['Scan Direction'].str.lower().isin(['rev', 'rv', 'reverse'])].sort_values('Filename')

                if fwd.empty or rev.empty:
                    continue

                has_any_pair = True
                n_pairs = min(len(fwd), len(rev))
                for i in range(n_pairs):
                    fw_eff = pd.to_numeric(fwd.iloc[i].get('Efficiency [.]'), errors='coerce')
                    rv_eff = pd.to_numeric(rev.iloc[i].get('Efficiency [.]'), errors='coerce')
                    if pd.isna(fw_eff):
                        fw_eff = 0.0
                    if pd.isna(rv_eff):
                        rv_eff = 0.0
                    avg = (fw_eff + rv_eff) / 2.0
                    if avg > best_avg:
                        best_avg = avg
                        best_pair = (fwd.index[i], rev.index[i])

            if best_pair is not None:
                kept.append(df.loc[[best_pair[0], best_pair[1]]])
            elif not has_any_pair:
                kept.append(sub_group)

        if not kept:
            return pd.DataFrame(columns=df.columns)
        return pd.concat(kept, ignore_index=True)

    def generate_plot(self):
        substrate_mode = hasattr(self, 'plot_by_substrate_var') and self.plot_by_substrate_var.get()
        if self.measurements_data.empty:
            messagebox.showwarning("Warning", "No data available for plotting.")
            return
        if not substrate_mode and self.conditions_data.empty:
            messagebox.showwarning("Warning", "No conditions assigned. Assign conditions or enable 'Plot by Substrate ID'.")
            return

        # Get selected parameters from all plot slots
        selected_params = []
        for slot in self.plot_slots:
            param = slot['combobox'].get()
            if param:
                selected_params.append(param)

        if not selected_params:
            messagebox.showwarning("Warning", "Please select at least one parameter to plot.")
            return

        # Use only the data currently displayed in the measurements table
        # Extract data from the treeview to reflect exactly what the user sees
        displayed_data = []
        for item in self.measurements_tree.get_children():
            values = self.measurements_tree.item(item, 'values')
            row_dict = {col: values[i] for i, col in enumerate(self.display_columns)}
            displayed_data.append(row_dict)

        if not displayed_data:
            messagebox.showwarning("Warning", "No data displayed in measurements table.")
            return

        displayed_df = pd.DataFrame(displayed_data)
        displayed_df = displayed_df.drop_duplicates(subset='Filename')

        # Convert numeric columns to proper numeric types
        numeric_columns = ['Voc [V]', 'Jsc [mA/cm2]', 'FF [.]', 'Efficiency [.]',
                          'Pmpp [W/m2]', 'Vmpp [V]', 'Jmpp [mA/cm2]', 'Roc [Ohm.m2]', 'Rsc [Ohm.m2]', 'Scan Speed [V/s]', 'Cell Area [cm2]']
        for col in numeric_columns:
            if col in displayed_df.columns:
                displayed_df[col] = pd.to_numeric(displayed_df[col], errors='coerce')

        # Prepare data for plotting
        if substrate_mode:
            # Each Substrate ID is its own condition
            plot_data = displayed_df.copy()
            plot_data['Condition'] = plot_data['Substrate ID']
        else:
            # Merge with conditions (deduplicate to avoid row doubling)
            conditions_dedup = self.conditions_data.drop_duplicates(subset='Substrate ID')
            plot_data = displayed_df.merge(
                conditions_dedup,
                on='Substrate ID',
                how='inner'
            )

        if plot_data.empty:
            messagebox.showwarning("Warning", "No data available after joining measurements with conditions.")
            return

        # Filter to only conditions selected in the condition filter listbox
        if hasattr(self, 'plot_cond_listbox') and self.plot_cond_listbox.size() > 0:
            selected_indices = self.plot_cond_listbox.curselection()
            if selected_indices:
                selected_conds = {self.plot_cond_listbox.get(i) for i in selected_indices}
                plot_data = plot_data[plot_data['Condition'].isin(selected_conds)]
            if plot_data.empty:
                messagebox.showwarning("Warning", "No data for the selected conditions.")
                return

        # Filter to best measurement per cell if checkbox is ticked
        if self.best_measurement_var.get():
            plot_data = self._filter_best_measurement(plot_data)
            if plot_data.empty:
                messagebox.showwarning("Warning", "No complete fwd/rev pairs found.")
                return

        # Clear previous plot
        for widget in self.condition_plot_frame.winfo_children():
            widget.destroy()

        # Determine grid layout
        n_plots = len(selected_params)
        if n_plots == 1:
            nrows, ncols = 1, 1
            figsize = (7, 6)
        elif n_plots == 2:
            nrows, ncols = 1, 2
            figsize = (14, 6)
        elif n_plots == 3:
            nrows, ncols = 2, 2
            figsize = (14, 12)
        else:  # 4 plots
            nrows, ncols = 2, 2
            figsize = (14, 12)

        # Create figure with subplots
        fig, axes = plt.subplots(nrows, ncols, figsize=figsize)

        # Flatten axes array for easier iteration
        if n_plots == 1:
            axes = [axes]
        else:
            axes = axes.flatten() if n_plots > 2 else axes

        # Get the desired condition order
        condition_order = self.get_condition_order(plot_data)

        # Get selected color palette or use manual colors
        palette_name = self.color_palette_combobox.get()

        # Check if we have manual colors defined and use them, otherwise use selected palette
        if self.manual_colors:
            # Use manual colors - create a palette dictionary for the conditions in order
            palette = {condition: self.manual_colors.get(condition, "#1f77b4") for condition in condition_order}
        else:
            palette = self.color_palettes[palette_name]

        # Create each subplot
        for i, param in enumerate(selected_params):
            ax = axes[i]
            slot = self.plot_slots[i]

            # Create box plot — avoid hue to prevent conflict with stripplot
            if isinstance(palette, dict):
                box_palette = [palette.get(c, "#1f77b4") for c in condition_order]
            elif palette is not None:
                try:
                    box_palette = sns.color_palette(palette, n_colors=len(condition_order))
                except Exception:
                    box_palette = None
            else:
                box_palette = None
            sns.boxplot(data=plot_data, x='Condition', y=param, ax=ax,
                        palette=box_palette, order=condition_order, legend=False, showfliers=False)

            # Plot data points manually with matplotlib to avoid seaborn doubling
            import numpy as np
            # Get the actual x tick positions from the boxplot axes
            tick_positions = {label.get_text(): pos for pos, label in
                              zip(ax.get_xticks(), ax.get_xticklabels())}
            # Fallback: if tick labels aren't set yet, use sequential positions
            if not tick_positions or all(t == '' for t in tick_positions):
                tick_positions = {c: idx for idx, c in enumerate(condition_order)}
            dir_colors = {'fwd': 'black', 'rev': 'red'}
            jitter_strength = 0.15
            plotted_dirs = set()  # Reset per subplot so each gets a legend
            for _, row in plot_data.iterrows():
                cond = row.get('Condition', '')
                if cond not in tick_positions:
                    continue
                x_pos = tick_positions[cond]
                y_val = pd.to_numeric(row.get(param), errors='coerce')
                if pd.isna(y_val):
                    continue
                scan_dir = str(row.get('Scan Direction', '')).lower().strip()
                color = dir_colors.get(scan_dir, 'gray')
                jitter = np.random.uniform(-jitter_strength, jitter_strength)
                dir_label = {'fwd': 'Forward', 'rev': 'Reverse'}.get(scan_dir, scan_dir)
                label = dir_label if scan_dir not in plotted_dirs else None
                ax.scatter(x_pos + jitter, y_val, color=color, s=12, alpha=0.7,
                           zorder=5, edgecolors='white', linewidths=0.3, label=label)
                if label:
                    plotted_dirs.add(scan_dir)

            # Set y-axis limits if provided for this specific plot
            try:
                ymin = slot['ymin'].get().strip()
                ymax = slot['ymax'].get().strip()

                if ymin and ymax:
                    ax.set_ylim(float(ymin), float(ymax))
                elif ymin:
                    ax.set_ylim(bottom=float(ymin))
                elif ymax:
                    ax.set_ylim(top=float(ymax))
            except ValueError:
                pass  # Use auto limits if invalid

            try:
                ylabel_size = int(self.ylabel_size_var.get())
            except (ValueError, AttributeError):
                ylabel_size = 20
            try:
                ytick_size = int(self.ytick_size_var.get())
            except (ValueError, AttributeError):
                ytick_size = 10
            try:
                xtick_size = int(self.xtick_size_var.get())
            except (ValueError, AttributeError):
                xtick_size = 16
            try:
                xtick_angle = int(self.xtick_angle_var.get())
            except (ValueError, AttributeError):
                xtick_angle = 45

            ax.set_ylabel(param, fontsize=ylabel_size)
            ax.set_xlabel('')  # Remove x-axis label completely

            # Add horizontal grid lines only
            ax.grid(True, axis='y', alpha=0.4, linestyle='--')

            # Set tick parameters with larger labels
            ax.tick_params(axis='y', labelsize=ytick_size)

            # X-axis tick labels with user-defined angle
            ha = 'right' if xtick_angle > 0 else 'center'
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=xtick_angle, ha=ha, fontsize=xtick_size)

            ax.legend(fontsize=9, loc='best')

        # Hide unused subplots
        for i in range(n_plots, nrows * ncols):
            axes[i].set_visible(False)

        # Tight layout to ensure everything fits
        plt.tight_layout()

        # Save the figure reference and plot data for later use
        self.current_figure = fig
        self.current_plot_data = plot_data.copy()

        # Embed plot in tkinter
        canvas = FigureCanvasTkAgg(fig, master=self.condition_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def save_plot(self):
        if self.current_figure is None:
            messagebox.showwarning("Warning", "No plot to save. Please generate a plot first.")
            return
        
        # Get the save file path
        file_path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[
                ("PNG files", "*.png"),
                ("JPEG files", "*.jpg"),
                ("PDF files", "*.pdf"),
                ("SVG files", "*.svg"),
                ("All files", "*.*")
            ],
            title="Save Plot As"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Save the figure
            self.current_figure.savefig(file_path, bbox_inches='tight', dpi=300)
            messagebox.showinfo("Success", f"Plot saved successfully to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save plot: {str(e)}")

    def export_condition_data(self):
        """Export the current condition plot data to Excel format"""
        if self.current_plot_data is None or self.current_plot_data.empty:
            messagebox.showwarning("Warning", "No plot data to export. Please generate a plot first.")
            return

        # Get the save file path
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ],
            title="Export Condition Data As"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Create a copy of the data for export
            export_data = self.current_plot_data.copy()

            # Get the condition order used in the plot
            condition_order = self.get_condition_order(export_data)

            # Sort data by condition order for better readability
            export_data['Condition_Order'] = export_data['Condition'].map(
                {cond: idx for idx, cond in enumerate(condition_order)}
            )
            export_data = export_data.sort_values(['Condition_Order', 'Substrate ID'])
            export_data = export_data.drop('Condition_Order', axis=1)

            # Select relevant columns for export (excluding internal columns)
            columns_to_export = [
                'Substrate ID', 'Condition', 'Pixel', 'Scan Direction',
                'Voc [V]', 'Jsc [mA/cm2]', 'FF [.]', 'Efficiency [.]',
                'Pmpp [W/m2]', 'Vmpp [V]', 'Jmpp [mA/cm2]',
                'Roc [Ohm.m2]', 'Rsc [Ohm.m2]', 'Scan Speed [V/s]', 'Cell Area [cm2]'
            ]

            # Filter to only include columns that exist in the data
            available_columns = [col for col in columns_to_export if col in export_data.columns]
            export_data_filtered = export_data[available_columns]

            # Export based on file extension
            if file_path.lower().endswith('.csv'):
                export_data_filtered.to_csv(file_path, index=False)
            else:
                # Default to Excel format (requires openpyxl: pip install openpyxl)
                export_data_filtered.to_excel(file_path, index=False, sheet_name='Condition Data')

            # Show success message with summary
            total_measurements = len(export_data_filtered)
            unique_conditions = export_data_filtered['Condition'].nunique()
            messagebox.showinfo(
                "Success",
                f"Data exported successfully to:\n{file_path}\n\n"
                f"Exported {total_measurements} measurements across {unique_conditions} conditions."
            )

        except ImportError as e:
            if "openpyxl" in str(e).lower():
                messagebox.showerror(
                    "Missing Dependency",
                    "Excel export requires the 'openpyxl' package.\n\n"
                    "Please install it using:\n"
                    "pip install openpyxl\n\n"
                    "Alternatively, you can export as CSV format."
                )
            else:
                messagebox.showerror("Import Error", f"Failed to export data: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {str(e)}")

    def setup_iv_plot(self, parent):
        # Create main container with two columns
        main_container = ttk.Frame(parent)
        main_container.pack(fill="both", expand=True, padx=10, pady=5)

        # Left side - Controls (takes up about 60% of width)
        left_frame = ttk.Frame(main_container)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Right side - Plot area (fixed square size)
        right_frame = ttk.Frame(main_container)
        right_frame.pack(side="right", fill="y", padx=(5, 0))

        # === LEFT SIDE CONTROLS ===

        # File controls
        control_frame = ttk.LabelFrame(left_frame, text="File Controls")
        control_frame.pack(fill="x", pady=(0, 5))

        button_row1 = ttk.Frame(control_frame)
        button_row1.pack(fill="x", padx=5, pady=5)

        browse_button = ttk.Button(button_row1, text="Load IV File", command=self.load_iv_data)
        browse_button.pack(side="left", padx=(0, 5))

        select_btn = ttk.Button(button_row1, text="Plot Selected", command=self.load_iv_from_selection)
        select_btn.pack(side="left", padx=5)

        save_btn = ttk.Button(button_row1, text="Save Plot", command=self.save_current_plot)
        save_btn.pack(side="left", padx=5)

        button_row2 = ttk.Frame(control_frame)
        button_row2.pack(fill="x", padx=5, pady=(0, 5))

        pair_btn = ttk.Button(button_row2, text="Pair FW↔RV (auto)", command=self.auto_pair_and_plot)
        pair_btn.pack(side="left", padx=(0, 5))

        pair_sel_btn = ttk.Button(button_row2, text="Pair FW↔RV (selection)", command=self.auto_pair_selection_and_plot)
        pair_sel_btn.pack(side="left", padx=5)

        # Selection list
        selection_frame = ttk.LabelFrame(left_frame, text="Measurement Selection")
        selection_frame.pack(fill="both", expand=True, pady=(0, 5))

        ttk.Label(selection_frame, text="Select from loaded measurements:").pack(anchor="w", padx=5, pady=(5, 2))

        list_frame = ttk.Frame(selection_frame)
        list_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        self.iv_listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, height=8)
        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.iv_listbox.yview)
        self.iv_listbox.configure(yscrollcommand=vsb.set)
        self.refresh_iv_selection()
        self.iv_listbox.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Sort controls
        sort_frame = ttk.Frame(selection_frame)
        sort_frame.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Label(sort_frame, text="Sort:").pack(side="left", padx=(0, 5))
        ttk.Button(sort_frame, text="Voc ↑", command=lambda: self.sort_iv_selection_by_voc(True)).pack(side="left", padx=2)
        ttk.Button(sort_frame, text="Voc ↓", command=lambda: self.sort_iv_selection_by_voc(False)).pack(side="left", padx=2)

        # Axis controls
        axes_frame = ttk.LabelFrame(left_frame, text="Plot Controls")
        axes_frame.pack(fill="x", pady=(0, 5))

        axes_row1 = ttk.Frame(axes_frame)
        axes_row1.pack(fill="x", padx=5, pady=5)
        ttk.Label(axes_row1, text="Voltage:").pack(side="left", padx=(0, 5))
        self.iv_xmin = tk.StringVar(); self.iv_xmax = tk.StringVar()
        ttk.Label(axes_row1, text="Min").pack(side="left"); ttk.Entry(axes_row1, textvariable=self.iv_xmin, width=8).pack(side="left", padx=3)
        ttk.Label(axes_row1, text="Max").pack(side="left"); ttk.Entry(axes_row1, textvariable=self.iv_xmax, width=8).pack(side="left", padx=3)

        axes_row2 = ttk.Frame(axes_frame)
        axes_row2.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Label(axes_row2, text="Current:").pack(side="left", padx=(0, 5))
        self.iv_ymin = tk.StringVar(); self.iv_ymax = tk.StringVar()
        ttk.Label(axes_row2, text="Min").pack(side="left"); ttk.Entry(axes_row2, textvariable=self.iv_ymin, width=10).pack(side="left", padx=3)
        ttk.Label(axes_row2, text="Max").pack(side="left"); ttk.Entry(axes_row2, textvariable=self.iv_ymax, width=10).pack(side="left", padx=3)

        axes_row3 = ttk.Frame(axes_frame)
        axes_row3.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Label(axes_row3, text="Area [cm²]:").pack(side="left", padx=(0, 5))
        self.iv_area = tk.StringVar(value="0.04")
        ttk.Entry(axes_row3, textvariable=self.iv_area, width=8).pack(side="left", padx=3)

        ttk.Button(axes_row3, text="Apply", command=self.apply_iv_axes).pack(side="left", padx=(10, 5))
        ttk.Button(axes_row3, text="Autoscale", command=self.reset_iv_axes).pack(side="left", padx=2)

        # === RIGHT SIDE PLOT ===

        # Create a fixed-size frame for the plot (square)
        plot_container = ttk.LabelFrame(right_frame, text="IV Plot")
        plot_container.pack(fill="both", expand=True)

        # Create plot frame with fixed dimensions to maintain square aspect
        self.iv_plot_frame = ttk.Frame(plot_container)
        self.iv_plot_frame.pack(padx=10, pady=10)

        # Set a minimum size for the plot container to keep it reasonably sized
        plot_container.configure(width=500, height=500)
    
    def load_iv_data(self):
        file_path = filedialog.askopenfilename(filetypes=[("IV Files", "*.iv"), ("All Files", "*.*")])
        if not file_path:
            return
    
        try:
            data = self.parse_iv_data_for_plot(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse IV file: {str(e)}")
        try:
            self.plot_iv_curve(data)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to plot IV file: {str(e)}")
            
    
    def parse_iv_data_for_plot(self, filepath):
        with open(filepath, 'r') as f:
            content = f.read()

        # Extract cell area (labeled as m2 but actually in cm2)
        cell_area_cm2 = 0.04  # Default fallback
        area_match = re.search(r'Cell size \[m2\]:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)', content)
        if area_match:
            try:
                # Despite the label saying [m2], the value is actually in cm²
                cell_area_cm2 = float(area_match.group(1))
            except:
                cell_area_cm2 = 0.04

        # Extract sweep metadata to compute scan speed (V/s)
        vstart = self.extract_value(content, r'Vstart:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        vend = self.extract_value(content, r'Vend:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        npoints = None
        m_np = re.search(r'Number of points:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)', content)
        if m_np:
            try:
                npoints = int(float(m_np.group(1)))
            except:
                npoints = None
        delay_s = self.extract_value(content, r'Delay \[s\]:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        it_s = self.extract_value(content, r'Integration time \[s\]:\s*([-\d\.]+E[+-]\d+|[-\d\.]+)')
        scan_speed = None
        try:
            if vstart is not None and vend is not None and npoints and npoints > 1 and (delay_s is not None) and (it_s is not None):
                step_time = (delay_s + it_s)
                total_time = (npoints - 1) * step_time
                if total_time > 0:
                    scan_speed = abs(vend - vstart) / total_time
        except Exception:
            scan_speed = None


        data_section = re.search(    r'% MEASURED IV FRLOOP DATA\s*\nV \(measured\) \[V\]\s+I \(measured\) \[A\].*?\n(.*)', content, re.DOTALL)

        if not data_section:
            raise ValueError("Could not find IV data in file.")

        data_lines = data_section.group(1).strip().split('\n')
        voltages, currents = [], []

        for line in data_lines:
            values = line.split('\t')
            if len(values) >= 2:
                try:
                    voltages.append(float(values[0]))
                    currents.append(float(values[1]))
                except ValueError:
                    continue

        return pd.DataFrame({'Voltage (V)': voltages, 'Current (A)': currents, 'Cell Area [cm2]': cell_area_cm2})

    def convert_to_current_density(self, current_data, area_cm2=None):
        """Convert current (A) to current density (mA/cm²) using provided area or UI area"""
        try:
            # If area is not provided, try to get from UI
            if area_cm2 is None:
                area_cm2 = float(self.iv_area.get())
            if area_cm2 <= 0:
                return current_data  # Return original if invalid area
            # Convert A to mA/cm²: (A × 1000) / area_cm²
            return current_data * 1000 / area_cm2
        except (ValueError, AttributeError, TypeError):
            return current_data  # Return original if area parsing fails
    
    def plot_iv_curve(self, data):
        for widget in self.iv_plot_frame.winfo_children():
            widget.destroy()

        fig, ax = plt.subplots(figsize=(6, 6))
        # Convert current to current density using area from file if available
        area_cm2 = data['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in data.columns else None
        current_density = self.convert_to_current_density(data['Current (A)'], area_cm2)
        ax.plot(data['Voltage (V)'], current_density, marker='o', linestyle='-')
        ax.set_title("IV Curve")
        ax.set_xlabel("Voltage (V)", fontsize=12)
        ax.set_ylabel("Current Density (mA/cm²)", fontsize=12)
        ax.grid(True, alpha=0.4)
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.iv_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()


    def refresh_iv_selection(self):
        """Fill the IV selection list from the first tab table, including conditions, respecting sort order."""
        try:
            if hasattr(self, 'iv_listbox'):
                self.iv_listbox.delete(0, tk.END)
                # Build condition map
                cond_map = {}
                try:
                    if hasattr(self, 'conditions_data') and not self.conditions_data.empty:
                        for _, r in self.conditions_data.iterrows():
                            cond_map[str(r.get('Substrate ID',''))] = r.get('Condition','')
                except Exception:
                    cond_map = {}
                if 'Filename' in self.measurements_data.columns:
                    df = self.measurements_data.copy()
                    df = self._apply_iv_sort(df)
                    for _, row in df.iterrows():
                        fn = row.get('Filename', '')
                        subid = str(row.get('Substrate ID', ''))
                        pix = row.get('Pixel', '')
                        sd = row.get('Scan Direction', '')
                        cond = cond_map.get(subid, '—')
                        disp = f"{fn}  |  {subid}  |  Pixel {pix}  |  {sd}  |  {cond}"
                        self.iv_listbox.insert(tk.END, disp)
        except Exception:
            pass


    def load_iv_from_selection(self):
        """Parse and plot IV curve(s) for the selected entries from the first tab."""
        try:
            indices = self.iv_listbox.curselection()
        except Exception:
            indices = []
        if not indices:
            messagebox.showinfo("Select file(s)", "Please choose one or more measurements from the list.")
            return
        
        filepaths = []
        for idx in indices:
            sel = self.iv_listbox.get(idx)
            filename = sel.split('  |')[0].strip()
            if 'Filepath' in self.measurements_data.columns:
                matches = self.measurements_data[self.measurements_data['Filename'] == filename]
                if not matches.empty:
                    fp = matches.iloc[0].get('Filepath', None)
                    if fp: filepaths.append(fp)
        if not filepaths:
            messagebox.showerror("Not found", "Could not locate file paths for the selections.")
            return
        
        curves = []
        for fp in filepaths:
            try:
                df = self.parse_iv_data_for_plot(fp)
                curves.append((fp, df))
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load {os.path.basename(fp)}: {e}")
        if not curves:
            return
        self.plot_iv_curves_overlaid(curves)

    def _split_fw_rv_if_present(self, df):
        """Split a single IV sweep into FW/RV if a voltage direction reversal is found.
        Returns a list of (label_suffix, sub_df). If no reversal, returns [("", df)]."""
        v = df['Voltage (V)'].values
        if len(v) < 3:
            return [("", df)]
        dir0 = 1 if v[1] >= v[0] else -1
        change_idx = None
        for i in range(2, len(v)):
            d = 1 if v[i] >= v[i-1] else -1
            if d != dir0:
                change_idx = i-1
                break
        if change_idx is None:
            return [("", df)]
        df1 = df.iloc[:change_idx+1].reset_index(drop=True)
        df2 = df.iloc[change_idx+1:].reset_index(drop=True)
        # Label segments
        lab1 = " (FW)" if df1['Voltage (V)'].iloc[-1] > df1['Voltage (V)'].iloc[0] else " (RV)"
        lab2 = " (FW)" if df2['Voltage (V)'].iloc[-1] > df2['Voltage (V)'].iloc[0] else " (RV)"
        return [(lab1, df1), (lab2, df2)]

        def plot_iv_curves_overlaid(self, curves):
            """Overlay arbitrary curves; if a single file contains FW+RV, use split; style as FW/RV with pairwise colors."""
            for w in self.iv_plot_frame.winfo_children():
                w.destroy()
            fig, ax = plt.subplots(figsize=(8, 6))

            # Build condition map
            cond_map = {}
            try:
                if hasattr(self, 'conditions_data') and not self.conditions_data.empty:
                    for _, r in self.conditions_data.iterrows():
                        cond_map[str(r.get('Substrate ID',''))] = r.get('Condition','')
            except Exception:
                cond_map = {}

            import os
            for i, (fp, df) in enumerate(curves):
                parts = self._split_fw_rv_if_present(df)
                subid = None
                try:
                    if 'Filepath' in self.measurements_data.columns:
                        match = self.measurements_data[self.measurements_data['Filepath'] == fp]
                        if not match.empty:
                            subid = str(match.iloc[0].get('Substrate ID',''))
                except Exception:
                    subid = None
                cond = cond_map.get(subid or '', '—')
                c_fw, c_rv = self._pair_colors(i)
                for suf, part in parts:
                    # crude check for RV segment
                    is_rv = ('RV' in suf) or ('rv' in suf)
                    label = f"{subid or ''} | {cond} | {'RV' if is_rv else 'FW'}"
                    if is_rv:
                        ax.plot(part['Voltage (V)'], part['Current (A)'], linestyle='--', color=c_rv, alpha=0.5, label=label)
                    else:
                        ax.plot(part['Voltage (V)'], part['Current (A)'], linestyle='-', color=c_fw, label=label)

            ax.set_title("IV Curves (overlay)")
            ax.set_xlabel("Voltage (V)")
            ax.set_ylabel("Current (A)")
            ax.grid(True)
            ax.legend()
            canvas = FigureCanvasTkAgg(fig, master=self.iv_plot_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def sort_iv_selection_by_voc(self, ascending=True):
        """Sort the selection listbox by Voc [V]."""
        try:
            self._iv_sort_key = ('voc', ascending)
        except Exception:
            self._iv_sort_key = ('voc', ascending)
        self.refresh_iv_selection()

    def _apply_iv_sort(self, df):
        """Internal: apply current sort to the measurements_data for the selection list."""
        key = getattr(self, '_iv_sort_key', None)
        if not key:
            return df
        kind, asc = key
        try:
            if kind == 'voc' and 'Voc [V]' in df.columns:
                # Convert to numeric to be safe
                dfx = df.copy()
                dfx['__voc__'] = pd.to_numeric(dfx['Voc [V]'], errors='coerce')
                dfx = dfx.sort_values(by='__voc__', ascending=asc, na_position='last')
                return dfx
        except Exception:
            return df
        return df

        def _get_color_cycle(self):
            import matplotlib.pyplot as plt
            return plt.rcParams.get('axes.prop_cycle', None).by_key().get('color', ['C0','C1','C2','C3','C4','C5']) if plt.rcParams.get('axes.prop_cycle', None) else ['C0','C1','C2','C3','C4','C5']

        def _pair_colors(self, idx):
            colors = self._get_color_cycle()
            base = colors[idx % len(colors)]
            return base, base

    def auto_pair_and_plot(self):
        """Pair FW/RV across ALL loaded data and plot (FW solid, RV dashed; RV alpha=0.5)."""
        needed = {'Substrate ID','Pixel','Scan Direction','Filepath','Filename'}
        if not needed.issubset(set(self.measurements_data.columns)):
            messagebox.showwarning("Missing info", "Required columns missing; please load files in the first tab.")
            return

        df = self.measurements_data.copy()
        df['__dir__'] = df['Scan Direction'].astype(str).str.lower().replace({'fw':'fwd','forward':'fwd','rev':'rev','reverse':'rev'})

        cond_map = {}
        try:
            if hasattr(self, 'conditions_data') and not self.conditions_data.empty:
                for _, r in self.conditions_data.iterrows():
                    cond_map[str(r.get('Substrate ID',''))] = r.get('Condition','')
        except Exception:
            cond_map = {}

        pairs = []
        for (subid, pix), g in df.groupby(['Substrate ID','Pixel']):
            fwd_rows = g[g['__dir__']=='fwd']
            rev_rows = g[g['__dir__']=='rev']
            if not fwd_rows.empty and not rev_rows.empty:
                pairs.append((fwd_rows.iloc[0], rev_rows.iloc[0]))
            else:
                def guess_dir(fn):
                    s = str(fn).lower()
                    if any(t in s for t in ['_fw','-fw',' fw','(fw)','_fwd','-fwd',' fwd']): return 'fwd'
                    if any(t in s for t in ['_rv','-rv',' rv','(rv)','_rev','-rev',' rev']): return 'rev'
                    return None
                g2 = g.copy()
                g2['__guess__'] = g2['Filename'].apply(guess_dir)
                fwd_rows = g2[g2['__guess__']=='fwd']
                rev_rows = g2[g2['__guess__']=='rev']
                if not fwd_rows.empty and not rev_rows.empty:
                    pairs.append((fwd_rows.iloc[0], rev_rows.iloc[0]))

        if not pairs:
            messagebox.showinfo("No pairs found", "No FW/RV pairs detected.")
            return

        for w in self.iv_plot_frame.winfo_children():
            w.destroy()
        fig, ax = plt.subplots(figsize=(6, 6))
        self._last_ax = ax
        self._maybe_apply_axes(ax)

        # Get color cycle
        try:
            colors = plt.rcParams['axes.prop_cycle'].by_key().get('color', [])
        except Exception:
            colors = []

        for i, (fw_row, rv_row) in enumerate(pairs):
            subid = str(fw_row['Substrate ID'])
            cond = cond_map.get(subid, '—')
            base_color = colors[i % len(colors)] if colors else None
            # FWD
            try:
                df_fw = self.parse_iv_data_for_plot(fw_row['Filepath'])
                area_fw = df_fw['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in df_fw.columns else None
                current_density_fw = self.convert_to_current_density(df_fw['Current (A)'], area_fw)
                ax.plot(df_fw['Voltage (V)'], current_density_fw, linestyle='-', color=base_color, label=f"{subid} | {cond} | FW")
            except Exception as e:
                messagebox.showerror("Error", f"Failed FW: {e}")
            # REV
            try:
                df_rv = self.parse_iv_data_for_plot(rv_row['Filepath'])
                area_rv = df_rv['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in df_rv.columns else None
                current_density_rv = self.convert_to_current_density(df_rv['Current (A)'], area_rv)
                ax.plot(df_rv['Voltage (V)'], current_density_rv, linestyle='--', color=base_color, alpha=0.5, label=f"{subid} | {cond} | RV")
            except Exception as e:
                messagebox.showerror("Error", f"Failed RV: {e}")

        ax.set_title("IV Curves")
        ax.set_xlabel("Voltage (V)", fontsize=12)
        ax.set_ylabel("Current Density (mA/cm²)", fontsize=12)
        ax.grid(True, alpha=0.4)
        ax.legend()
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.iv_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()
    def auto_pair_selection_and_plot(self):
        """Pair FW/RV ONLY for selected items."""
        needed = {'Substrate ID','Pixel','Scan Direction','Filepath','Filename'}
        if not needed.issubset(set(self.measurements_data.columns)):
            messagebox.showwarning("Missing info", "Required columns missing; please load files in the first tab.")
            return

        try:
            indices = self.iv_listbox.curselection()
        except Exception:
            indices = []
        if not indices:
            messagebox.showinfo("Select file(s)", "Please select at least one measurement in the list.")
            return

        df = self.measurements_data.copy()
        df['__dir__'] = df['Scan Direction'].astype(str).str.lower().replace({'fw':'fwd','forward':'fwd','rev':'rev','reverse':'rev'})

        cond_map = {}
        try:
            if hasattr(self, 'conditions_data') and not self.conditions_data.empty:
                for _, r in self.conditions_data.iterrows():
                    cond_map[str(r.get('Substrate ID',''))] = r.get('Condition','')
        except Exception:
            cond_map = {}

        def guess_dir(fn):
            s = str(fn).lower()
            if any(t in s for t in ['_fw','-fw',' fw','(fw)','_fwd','-fwd',' fwd']): return 'fwd'
            if any(t in s for t in ['_rv','-rv',' rv','(rv)','_rev','-rev',' rev']): return 'rev'
            return None

        def rows_for(subid, pix, direction):
            g = df[(df['Substrate ID']==subid) & (df['Pixel']==pix)]
            g1 = g[g['__dir__']==direction]
            if not g1.empty:
                return g1
            g2 = g.copy()
            g2['__guess__'] = g2['Filename'].apply(guess_dir)
            return g2[g2['__guess__']==direction]

        selected_rows = []
        for idx in indices:
            disp = self.iv_listbox.get(idx)
            fname = disp.split('  |')[0].strip()
            r = df[df['Filename']==fname]
            if not r.empty:
                selected_rows.append(r.iloc[0])

        pairs = []
        for r in selected_rows:
            subid, pix = r['Substrate ID'], r['Pixel']
            d = r['__dir__']
            if d not in ['fwd','rev']:
                d = guess_dir(r['Filename']) or 'fwd'
            other = 'rev' if d=='fwd' else 'fwd'
            cand = rows_for(subid, pix, other)
            if not cand.empty:
                o = cand.iloc[0]
                if d=='fwd':
                    pairs.append((r, o))
                else:
                    pairs.append((o, r))

        if not pairs:
            messagebox.showinfo("No pairs found", "No FW/RV pairs found for the selection.")
            return

        for w in self.iv_plot_frame.winfo_children():
            w.destroy()
        fig, ax = plt.subplots(figsize=(6, 6))
        self._last_ax = ax
        self._maybe_apply_axes(ax)

        try:
            colors = plt.rcParams['axes.prop_cycle'].by_key().get('color', [])
        except Exception:
            colors = []

        for i, (fw_row, rv_row) in enumerate(pairs):
            subid = str(fw_row['Substrate ID'])
            cond = cond_map.get(subid, '—')
            base_color = colors[i % len(colors)] if colors else None
            try:
                df_fw = self.parse_iv_data_for_plot(fw_row['Filepath'])
                area_fw = df_fw['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in df_fw.columns else None
                current_density_fw = self.convert_to_current_density(df_fw['Current (A)'], area_fw)
                ax.plot(df_fw['Voltage (V)'], current_density_fw, linestyle='-', color=base_color, label=f"{subid} | {cond} | FW")
            except Exception as e:
                messagebox.showerror("Error", f"Failed FW: {e}")
            try:
                df_rv = self.parse_iv_data_for_plot(rv_row['Filepath'])
                area_rv = df_rv['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in df_rv.columns else None
                current_density_rv = self.convert_to_current_density(df_rv['Current (A)'], area_rv)
                ax.plot(df_rv['Voltage (V)'], current_density_rv, linestyle='--', color=base_color, alpha=0.5, label=f"{subid} | {cond} | RV")
            except Exception as e:
                messagebox.showerror("Error", f"Failed RV: {e}")

        ax.set_title("IV Curves")
        ax.set_xlabel("Voltage (V)", fontsize=12)
        ax.set_ylabel("Current Density (mA/cm²)", fontsize=12)
        ax.grid(True,alpha=0.4)
        ax.legend()
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.iv_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()
    def plot_iv_curves_overlaid(self, curves):
        """Overlay arbitrary curves; if a single file contains FW+RV, split; FW solid, RV dashed, RV alpha=0.5."""
        for w in self.iv_plot_frame.winfo_children():
            w.destroy()
        fig, ax = plt.subplots(figsize=(6, 6))
        self._last_ax = ax
        self._maybe_apply_axes(ax)

        # Build condition map
        cond_map = {}
        try:
            if hasattr(self, 'conditions_data') and not self.conditions_data.empty:
                for _, r in self.conditions_data.iterrows():
                    cond_map[str(r.get('Substrate ID',''))] = r.get('Condition','')
        except Exception:
            cond_map = {}

        import os
        try:
            colors = plt.rcParams['axes.prop_cycle'].by_key().get('color', [])
        except Exception:
            colors = []
        for i, (fp, df) in enumerate(curves):
            parts = self._split_fw_rv_if_present(df)
            subid = None
            try:
                if 'Filepath' in self.measurements_data.columns:
                    match = self.measurements_data[self.measurements_data['Filepath'] == fp]
                    if not match.empty:
                        subid = str(match.iloc[0].get('Substrate ID',''))
            except Exception:
                subid = None
            cond = cond_map.get(subid or '', '—')
            base_color = colors[i % len(colors)] if colors else None
            for suf, part in parts:
                is_rv = ('RV' in suf) or ('rv' in suf)
                label = f"{subid or ''} | {cond} | {'RV' if is_rv else 'FW'}"
                area_cm2 = part['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in part.columns else None
                current_density = self.convert_to_current_density(part['Current (A)'], area_cm2)
                if is_rv:
                    ax.plot(part['Voltage (V)'], current_density, linestyle='--', color=base_color, alpha=0.5, label=label)
                else:
                    ax.plot(part['Voltage (V)'], current_density, linestyle='-', color=base_color, label=label)

        ax.set_title("IV Curves (overlay)")
        ax.set_xlabel("Voltage (V)", fontsize=12)
        ax.set_ylabel("Current Density (mA/cm²)", fontsize=12)
        ax.grid(True, alpha=0.4)
        ax.legend()
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.iv_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def _maybe_apply_axes(self, ax):
        try:
            xmin = self.iv_xmin.get().strip() if hasattr(self, 'iv_xmin') else ''
            xmax = self.iv_xmax.get().strip() if hasattr(self, 'iv_xmax') else ''
            ymin = self.iv_ymin.get().strip() if hasattr(self, 'iv_ymin') else ''
            ymax = self.iv_ymax.get().strip() if hasattr(self, 'iv_ymax') else ''
        except Exception:
            xmin = xmax = ymin = ymax = ''
        def to_float(s):
            try: return float(s)
            except: return None
        x0, x1 = to_float(xmin), to_float(xmax)
        y0, y1 = to_float(ymin), to_float(ymax)
        if x0 is not None or x1 is not None:
            cur = list(ax.get_xlim())
            if x0 is not None: cur[0] = x0
            if x1 is not None: cur[1] = x1
            ax.set_xlim(cur)
        if y0 is not None or y1 is not None:
            cur = list(ax.get_ylim())
            if y0 is not None: cur[0] = y0
            if y1 is not None: cur[1] = y1
            ax.set_ylim(cur)

    def apply_iv_axes(self):
        try:
            if hasattr(self, '_last_ax') and self._last_ax:
                self._maybe_apply_axes(self._last_ax)
                self._last_ax.figure.canvas.draw_idle()
        except Exception:
            pass

    def reset_iv_axes(self):
        try:
            if hasattr(self, 'iv_xmin'): self.iv_xmin.set('')
            if hasattr(self, 'iv_xmax'): self.iv_xmax.set('')
            if hasattr(self, 'iv_ymin'): self.iv_ymin.set('')
            if hasattr(self, 'iv_ymax'): self.iv_ymax.set('')
        except Exception:
            pass
        try:
            if hasattr(self, '_last_ax') and self._last_ax:
                self._last_ax.relim(); self._last_ax.autoscale(); self._last_ax.figure.canvas.draw_idle()
        except Exception:
            pass


    def save_current_plot(self):
        """Save the current IV plot to PNG/SVG/PDF."""
        try:
            ax = getattr(self, '_last_ax', None)
            fig = ax.figure if ax is not None else None
        except Exception:
            fig = None
        if fig is None:
            messagebox.showinfo("No plot", "There is no plot to save yet. Create a plot first.")
            return
        try:
            fpath = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG Image","*.png"), ("SVG Vector","*.svg"), ("PDF Document","*.pdf")],
                initialfile="IV_plot.png",
                title="Save IV Plot"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Save dialog: {e}")
            return
        if not fpath:
            return
        try:
            fig.savefig(fpath, dpi=300, bbox_inches="tight")
            messagebox.showinfo("Saved", f"Plot saved to:\n{fpath}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save plot: {e}")

    def setup_best_pce_tab(self, parent):
        """Setup the Best PCE tab with a table and interactive JV plot."""
        # Main horizontal split: left = table, right = plot
        main_container = ttk.PanedWindow(parent, orient=tk.HORIZONTAL)
        main_container.pack(fill="both", expand=True, padx=10, pady=5)

        # Left side: controls + table
        left_frame = ttk.Frame(main_container)
        main_container.add(left_frame, weight=1)

        # Refresh button
        control_frame = ttk.Frame(left_frame)
        control_frame.pack(fill="x", pady=(0, 5))

        ttk.Button(control_frame, text="Refresh Best PCE", command=self.compute_best_pce).pack(side="left", padx=5, pady=5)
        ttk.Button(control_frame, text="Export to Excel", command=self.export_best_pce).pack(side="left", padx=5, pady=5)
        ttk.Button(control_frame, text="Export to Existing Excel", command=self.export_best_pce_to_existing).pack(side="left", padx=5, pady=5)

        # Table
        table_frame = ttk.LabelFrame(left_frame, text="Best PCE per Cell (click row to plot JV)")
        table_frame.pack(fill="both", expand=True)

        best_pce_columns = ('Condition', 'Substrate ID', 'Pixel', 'Avg PCE [%]', 'Fwd PCE [%]', 'Rev PCE [%]',
                            'Avg Voc [V]', 'Avg Jsc [mA/cm2]', 'Avg FF [%]')
        self.best_pce_tree = ttk.Treeview(table_frame, columns=best_pce_columns, show='headings')

        for col in best_pce_columns:
            self.best_pce_tree.heading(col, text=col)
            width = 140 if col == 'Condition' else (120 if col == 'Substrate ID' else 100)
            self.best_pce_tree.column(col, width=width, anchor='center')

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.best_pce_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.best_pce_tree.xview)
        self.best_pce_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.best_pce_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Bind row selection to plot
        self.best_pce_tree.bind('<<TreeviewSelect>>', self.on_best_pce_row_click)

        # Store the best-pce pair data for plotting
        self.best_pce_pairs = {}  # {substrate_id: {'fwd_filepath': ..., 'rev_filepath': ...}}

        # Right side: JV plot
        right_frame = ttk.Frame(main_container)
        main_container.add(right_frame, weight=1)

        plot_label_frame = ttk.LabelFrame(right_frame, text="JV Curve")
        plot_label_frame.pack(fill="both", expand=True)

        self.best_pce_plot_frame = ttk.Frame(plot_label_frame)
        self.best_pce_plot_frame.pack(fill="both", expand=True, padx=5, pady=5)

    def compute_best_pce(self):
        """Compute best averaged PCE (fwd+rev) per condition (or per Substrate ID if no condition assigned).

        When multiple Substrate IDs share the same condition, they are pooled
        together and the single best fwd/rev pair across all of them is kept.
        """
        if self.measurements_data.empty:
            messagebox.showwarning("Warning", "No measurement data loaded.")
            return

        df = self.measurements_data.copy()
        df['Efficiency [.]'] = pd.to_numeric(df['Efficiency [.]'], errors='coerce')
        df['Voc [V]'] = pd.to_numeric(df['Voc [V]'], errors='coerce')
        df['Jsc [mA/cm2]'] = pd.to_numeric(df['Jsc [mA/cm2]'], errors='coerce')
        df['FF [.]'] = pd.to_numeric(df['FF [.]'], errors='coerce')

        # Normalize scan direction
        df['__dir__'] = df['Scan Direction'].astype(str).str.lower().str.strip()
        df['__dir__'] = df['__dir__'].replace({'fw': 'fwd', 'forward': 'fwd', 'reverse': 'rev', 'rv': 'rev'})

        # Merge with conditions — use Condition as grouping key when available
        if not self.conditions_data.empty:
            df = df.merge(self.conditions_data[['Substrate ID', 'Condition']], on='Substrate ID', how='left')
        else:
            df['Condition'] = None

        # For substrates without a condition, use Substrate ID as its own group
        df['__group__'] = df['Condition'].fillna(df['Substrate ID'])

        results = []
        self.best_pce_pairs = {}

        for group_key, group_df in df.groupby('__group__'):
            best_avg_pce = -float('inf')
            best_row = None

            for (subid, pixel), pix_group in group_df.groupby(['Substrate ID', 'Pixel']):
                fwd_rows = pix_group[pix_group['__dir__'] == 'fwd'].sort_values('Filename')
                rev_rows = pix_group[pix_group['__dir__'] == 'rev'].sort_values('Filename')

                if fwd_rows.empty or rev_rows.empty:
                    continue

                n_pairs = min(len(fwd_rows), len(rev_rows))
                for i in range(n_pairs):
                    fw = fwd_rows.iloc[i]
                    rv = rev_rows.iloc[i]
                    fw_eff = fw['Efficiency [.]'] if not pd.isna(fw['Efficiency [.]']) else 0.0
                    rv_eff = rv['Efficiency [.]'] if not pd.isna(rv['Efficiency [.]']) else 0.0
                    avg_pce = (fw_eff + rv_eff) / 2.0

                    if avg_pce > best_avg_pce:
                        best_avg_pce = avg_pce
                        fw_voc = fw['Voc [V]'] if not pd.isna(fw['Voc [V]']) else 0.0
                        rv_voc = rv['Voc [V]'] if not pd.isna(rv['Voc [V]']) else 0.0
                        fw_jsc = fw['Jsc [mA/cm2]'] if not pd.isna(fw['Jsc [mA/cm2]']) else 0.0
                        rv_jsc = rv['Jsc [mA/cm2]'] if not pd.isna(rv['Jsc [mA/cm2]']) else 0.0
                        fw_ff = fw['FF [.]'] if not pd.isna(fw['FF [.]']) else 0.0
                        rv_ff = rv['FF [.]'] if not pd.isna(rv['FF [.]']) else 0.0
                        condition_label = fw.get('Condition', '') or ''
                        best_row = {
                            'Condition': condition_label if condition_label else '—',
                            'Substrate ID': subid,
                            'Pixel': pixel,
                            'Avg PCE [%]': round(avg_pce, 3),
                            'Fwd PCE [%]': round(fw_eff, 3),
                            'Rev PCE [%]': round(rv_eff, 3),
                            'Avg Voc [V]': round((fw_voc + rv_voc) / 2.0, 4),
                            'Avg Jsc [mA/cm2]': round((fw_jsc + rv_jsc) / 2.0, 3),
                            'Avg FF [%]': round((fw_ff + rv_ff) / 2.0, 2),
                            'fwd_filepath': fw.get('Filepath', ''),
                            'rev_filepath': rv.get('Filepath', ''),
                            '__group__': group_key,
                            # Individual fwd/rev params for plot annotation
                            'Fwd Voc [V]': round(fw_voc, 4),
                            'Rev Voc [V]': round(rv_voc, 4),
                            'Fwd Jsc [mA/cm2]': round(fw_jsc, 3),
                            'Rev Jsc [mA/cm2]': round(rv_jsc, 3),
                            'Fwd FF [%]': round(fw_ff, 2),
                            'Rev FF [%]': round(rv_ff, 2),
                        }

            if best_row is not None:
                results.append(best_row)
                self.best_pce_pairs[group_key] = {
                    'fwd_filepath': best_row['fwd_filepath'],
                    'rev_filepath': best_row['rev_filepath'],
                    'params': best_row,
                }

        # Sort by Avg PCE descending
        results.sort(key=lambda x: x['Avg PCE [%]'], reverse=True)

        # Populate tree
        for item in self.best_pce_tree.get_children():
            self.best_pce_tree.delete(item)

        display_cols = ('Condition', 'Substrate ID', 'Pixel', 'Avg PCE [%]', 'Fwd PCE [%]', 'Rev PCE [%]',
                        'Avg Voc [V]', 'Avg Jsc [mA/cm2]', 'Avg FF [%]')
        for row in results:
            values = [row[col] for col in display_cols]
            self.best_pce_tree.insert('', 'end', values=values)

        # Store for export
        self._best_pce_results = results

        if not results:
            messagebox.showinfo("Info", "No complete fwd/rev pairs found for any cell.")

    def on_best_pce_row_click(self, event=None):
        """Plot JV curves for the selected row in the Best PCE table."""
        selected = self.best_pce_tree.selection()
        if not selected:
            return

        item_values = self.best_pce_tree.item(selected[0], 'values')
        # Columns: Condition, Substrate ID, Pixel, Avg PCE, Fwd PCE, Rev PCE, ...
        condition = item_values[0]
        subid = item_values[1]
        pixel = item_values[2]
        avg_pce = item_values[3]

        # Look up pair by condition (group key), fall back to substrate ID
        pair = self.best_pce_pairs.get(condition) or self.best_pce_pairs.get(subid)
        if not pair:
            return

        fwd_fp = pair['fwd_filepath']
        rev_fp = pair['rev_filepath']
        p = pair.get('params', {})

        # Clear previous plot
        for w in self.best_pce_plot_frame.winfo_children():
            w.destroy()

        fig, ax = plt.subplots(figsize=(6, 5))

        try:
            colors = plt.rcParams['axes.prop_cycle'].by_key().get('color', ['#1f77b4', '#ff7f0e'])
        except Exception:
            colors = ['#1f77b4', '#ff7f0e']

        # Plot forward
        try:
            df_fw = self.parse_iv_data_for_plot(fwd_fp)
            area_fw = df_fw['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in df_fw.columns else None
            jd_fw = self.convert_to_current_density(df_fw['Current (A)'], area_fw)
            ax.plot(df_fw['Voltage (V)'], jd_fw, linestyle='-', color=colors[0], label="Forward")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to plot forward curve: {e}")

        # Plot reverse
        try:
            df_rv = self.parse_iv_data_for_plot(rev_fp)
            area_rv = df_rv['Cell Area [cm2]'].iloc[0] if 'Cell Area [cm2]' in df_rv.columns else None
            jd_rv = self.convert_to_current_density(df_rv['Current (A)'], area_rv)
            ax.plot(df_rv['Voltage (V)'], jd_rv, linestyle='--', color=colors[1], alpha=0.7, label="Reverse")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to plot reverse curve: {e}")

        title_label = condition if condition != '—' else subid
        ax.set_title(f"{title_label} — {subid} Pixel {pixel} — Avg PCE: {avg_pce}%", fontsize=11)
        ax.set_xlabel("Voltage (V)", fontsize=12)
        ax.set_ylabel("Current Density (mA/cm²)", fontsize=12)
        ax.set_xlim(0, 1.2)
        ax.set_ylim(-22, 5)
        ax.grid(True, alpha=0.4)
        ax.legend(fontsize=9, loc='upper left')

        # Build parameter table as text annotation
        param_text = (
            f"{'':8s} {'Fwd':>8s}  {'Rev':>8s}\n"
            f"{'PCE [%]':8s} {p.get('Fwd PCE [%]',''):>8}  {p.get('Rev PCE [%]',''):>8}\n"
            f"{'Voc [V]':8s} {p.get('Fwd Voc [V]',''):>8}  {p.get('Rev Voc [V]',''):>8}\n"
            f"{'Jsc':8s} {p.get('Fwd Jsc [mA/cm2]',''):>8}  {p.get('Rev Jsc [mA/cm2]',''):>8}\n"
            f"{'FF [%]':8s} {p.get('Fwd FF [%]',''):>8}  {p.get('Rev FF [%]',''):>8}"
        )
        ax.text(0.98, 0.02, param_text, transform=ax.transAxes,
                fontsize=8, fontfamily='monospace', verticalalignment='bottom',
                horizontalalignment='right',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='wheat', alpha=0.8))

        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.best_pce_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def export_best_pce(self):
        """Export best PCE table to Excel/CSV."""
        if not hasattr(self, '_best_pce_results') or not self._best_pce_results:
            messagebox.showwarning("Warning", "No best PCE data to export. Click 'Refresh Best PCE' first.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export Best PCE Data"
        )
        if not file_path:
            return

        try:
            export_cols = ['Condition', 'Substrate ID', 'Pixel', 'Avg PCE [%]', 'Fwd PCE [%]', 'Rev PCE [%]',
                           'Avg Voc [V]', 'Avg Jsc [mA/cm2]', 'Avg FF [%]']
            export_df = pd.DataFrame(self._best_pce_results)[export_cols]

            if file_path.lower().endswith('.csv'):
                export_df.to_csv(file_path, index=False)
            else:
                export_df.to_excel(file_path, index=False, sheet_name='Best PCE')

            messagebox.showinfo("Success", f"Exported {len(export_df)} cells to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")


    def export_best_pce_to_existing(self):
        """Open an existing Excel file, match Substrate IDs, and add Avg PCE column."""
        if not hasattr(self, '_best_pce_results') or not self._best_pce_results:
            messagebox.showwarning("Warning", "No best PCE data to export. Click 'Refresh Best PCE' first.")
            return

        # Ask user to pick existing Excel file
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Select Existing Excel File to Add PCE Data"
        )
        if not file_path:
            return

        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Missing Dependency",
                                 "This feature requires 'openpyxl'.\n\nInstall with: pip install openpyxl")
            return

        try:
            # Read the existing Excel file
            existing_df = pd.read_excel(file_path, engine='openpyxl')

            # Find the column that contains sample/substrate IDs
            # Try common column names
            id_column = None
            for col in existing_df.columns:
                col_lower = str(col).lower().strip()
                if col_lower in ('substrate id', 'substrate_id', 'sample id', 'sample_id',
                                 'deposition id', 'deposition_id', 'id', 'sample', 'substrate'):
                    id_column = col
                    break

            if id_column is None:
                # Show a dialog letting the user pick the column
                col_picker = tk.Toplevel(self.root)
                col_picker.title("Select ID Column")
                col_picker.geometry("350x200")
                col_picker.grab_set()

                ttk.Label(col_picker, text="Which column contains the Sample/Substrate IDs?").pack(padx=10, pady=10)

                col_var = tk.StringVar()
                col_combo = ttk.Combobox(col_picker, values=list(existing_df.columns), textvariable=col_var,
                                         state="readonly", width=40)
                col_combo.pack(padx=10, pady=5)
                if len(existing_df.columns) > 0:
                    col_combo.current(0)

                result = {'column': None}

                def on_ok():
                    result['column'] = col_var.get()
                    col_picker.destroy()

                ttk.Button(col_picker, text="OK", command=on_ok).pack(pady=10)
                self.root.wait_window(col_picker)

                id_column = result['column']
                if not id_column:
                    return

            # Compute best avg PCE per Substrate ID (not per condition group)
            # so every substrate gets its own value for matching
            meas_df = self.measurements_data.copy()
            meas_df['Efficiency [.]'] = pd.to_numeric(meas_df['Efficiency [.]'], errors='coerce')
            meas_df['__dir__'] = meas_df['Scan Direction'].astype(str).str.lower().str.strip()
            meas_df['__dir__'] = meas_df['__dir__'].replace({'fw': 'fwd', 'forward': 'fwd', 'reverse': 'rev', 'rv': 'rev'})

            pce_map = {}
            for subid, sub_group in meas_df.groupby('Substrate ID'):
                best_avg = -float('inf')
                for pixel, pix_group in sub_group.groupby('Pixel'):
                    fwd_rows = pix_group[pix_group['__dir__'] == 'fwd'].sort_values('Filename')
                    rev_rows = pix_group[pix_group['__dir__'] == 'rev'].sort_values('Filename')
                    if fwd_rows.empty or rev_rows.empty:
                        continue
                    n_pairs = min(len(fwd_rows), len(rev_rows))
                    for i in range(n_pairs):
                        fw_eff = fwd_rows.iloc[i]['Efficiency [.]']
                        rv_eff = rev_rows.iloc[i]['Efficiency [.]']
                        fw_eff = fw_eff if not pd.isna(fw_eff) else 0.0
                        rv_eff = rv_eff if not pd.isna(rv_eff) else 0.0
                        avg = (fw_eff + rv_eff) / 2.0
                        if avg > best_avg:
                            best_avg = avg
                if best_avg > -float('inf'):
                    pce_map[str(subid).strip()] = round(best_avg, 3)

            # Normalize ID strings: strip whitespace and remove ".0" float suffix
            def normalize_id(val):
                s = str(val).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                return s

            # Normalize pce_map keys
            pce_map = {normalize_id(k): v for k, v in pce_map.items()}

            # Match Substrate IDs and build the new column values
            id_values = existing_df[id_column].apply(normalize_id)
            pce_values = id_values.map(pce_map)

            matched = pce_values.notna().sum()
            total = len(existing_df)

            # Check for unmatched IDs and warn before writing
            unmatched_excel = sorted(set(id_values[pce_values.isna()].unique()) - {'', 'nan', 'None'})
            if unmatched_excel:
                id_list = ', '.join(unmatched_excel)
                proceed = messagebox.askyesno(
                    "Unmatched Sample IDs",
                    f"Matched {matched}/{total} rows.\n\n"
                    f"No Malibu measurement data for ID(s): {id_list}\n\n"
                    f"These cells will be left empty. Continue with export?"
                )
                if not proceed:
                    return

            # Use openpyxl directly to add/overwrite a single column,
            # preserving all existing sheets, formatting, etc.
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Find or create the target column
            col_name = 'Avg PCE [%] (Malibu)'
            target_col = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=c).value == col_name:
                    target_col = c
                    break
            if target_col is None:
                target_col = ws.max_column + 1
                ws.cell(row=1, column=target_col, value=col_name)

            # Write the PCE values
            for i, val in enumerate(pce_values):
                cell = ws.cell(row=i + 2, column=target_col)
                if pd.notna(val):
                    cell.value = val
                else:
                    cell.value = None

            wb.save(file_path)
            wb.close()

            messagebox.showinfo("Success",
                                f"Added '{col_name}' column to:\n{file_path}\n\n"
                                f"Matched {matched}/{total} rows.")

        except PermissionError:
            messagebox.showerror("File In Use",
                                 "Cannot write to the file — it may be open in Excel.\n\n"
                                 "Please close the file and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = IVDataAnalyzer(root)
    root.mainloop()